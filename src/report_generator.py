# src/report_generator.py

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color
from openpyxl.chart import LineChart, ScatterChart, Reference, Series, BarChart
from openpyxl.drawing.image import Image as XLImage  # Rename to avoid conflict
import seaborn as sns
import matplotlib.pyplot as plt
from typing import Dict, List, Tuple
import os
import tempfile
from datetime import datetime, timedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph as RLParagraph,  # Rename to avoid conflict
    Spacer, Table, TableStyle, Image,
    PageBreak, KeepTogether
)
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.text import Paragraph as XLParagraph  # Excel Paragraph if needed
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties, RegularTextRun
import traceback

# Import MarketAnalyzer
from market_analyzer import MarketAnalyzer

class ReportGenerator:
    def __init__(self, 
                 market_analyzer: MarketAnalyzer, 
                 pricing_strategy: Dict, 
                 simulation_results: Dict):
        self.market_analyzer = market_analyzer
        self.pricing_strategy = pricing_strategy
        self.simulation_results = simulation_results
        self.wb = Workbook()
        self.setup_styles()
        
        # Store analysis results and verify supply data
        self.analysis_results = self.market_analyzer.analyze_market()
        
        # Debug print
        supply_data = self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']
        print("\nSupply Data Verification:")
        print(f"Active Units: {supply_data['active_units']}")
        print(f"Sold Out Units: {supply_data['sold_units']}")
        print(f"Standing Units: {supply_data['standing_units']}")
        print(f"Total Units: {supply_data['total_units']}")
        
    def setup_styles(self):
        """Setup consistent styling for the report"""
        self.styles = {
            'header': {
                'font': Font(name='Calibri', size=14, bold=True, color='000000'),
                'fill': PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
                'border': Border(bottom=Side(style='medium'))
            },
            'subheader': {
                'font': Font(name='Calibri', size=12, bold=True),
                'fill': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            },
            'metric_label': {
                'font': Font(name='Calibri', size=11),
                'alignment': Alignment(horizontal='left')
            },
            'metric_value': {
                'font': Font(name='Calibri', size=11, bold=True),
                'alignment': Alignment(horizontal='right')
            },
            'risk_high': PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid'),
            'risk_medium': PatternFill(start_color='FFE4B6', end_color='FFE4B6', fill_type='solid'),
            'risk_low': PatternFill(start_color='B6FFB6', end_color='B6FFB6', fill_type='solid')
        }
        
    def generate_report(self, output_file: str) -> None:
        """Generate complete Excel report"""
        self.temp_dir = tempfile.mkdtemp()
        
        try:
            # Create sheets
            self._create_executive_summary()
            
            # Create revenue analysis sheets for each absorption target
            self._add_revenue_analysis(
                self.wb.create_sheet("Revenue - 3M (50%)"), 
                target_period='3_month',
                target_absorption=50
            )
            self._add_revenue_analysis(
                self.wb.create_sheet("Revenue - 12M (65%)"), 
                target_period='12_month',
                target_absorption=65
            )
            self._add_revenue_analysis(
                self.wb.create_sheet("Revenue - 24M (82.5%)"), 
                target_period='24_month',
                target_absorption=82.5
            )
            self._add_revenue_analysis(
                self.wb.create_sheet("Revenue - 36M (100%)"), 
                target_period='36_month',
                target_absorption=100
            )
            
            self._create_unit_type_analysis()
            self._create_market_analysis()
            self._create_sensitivity_analysis()
            self._create_risk_analysis()
            self._create_recommendations()
            
            self.wb.save(output_file)
            
        finally:
            # Clean up temporary files
            for file in os.listdir(self.temp_dir):
                os.remove(os.path.join(self.temp_dir, file))
            os.rmdir(self.temp_dir)

    def _create_executive_summary(self) -> None:
        """Create executive summary dashboard"""
        ws = self.wb.active
        ws.title = "Executive Summary"
        
        # Title
        ws['A1'] = 'Surrey Market Analysis - Executive Summary'
        self._apply_style(ws['A1'], self.styles['header'])
        
        # Key Performance Indicators
        self._add_kpi_section(ws)
        
        # Market Overview Charts
        self._add_market_overview_charts(ws)
        
        # Unit Type Summary
        self._add_unit_type_summary(ws)
        
        # Key Risks and Opportunities
        self._add_risk_opportunity_summary(ws)

    def _create_unit_type_analysis(self) -> None:
        """Create detailed unit type analysis sheet"""
        ws = self.wb.create_sheet("Unit Type Analysis")
        
        # Add competitive analysis (which includes unit type analysis)
        self._add_competitive_analysis(ws)
        
        # Add unit type charts
        self._add_unit_type_charts(ws)
        
        # Add unit type recommendations
        self._add_unit_type_recommendations(ws)

    def _add_unit_type_metrics(self, ws) -> None:
        """Add detailed metrics for each unit type"""
        try:
            # Get project data from Surrey_Concrete_Launches_correct.csv
            project_data = {}
            active_section = False
            headers = None
            
            with open('data/raw/Surrey_Concrete_Launches_correct.csv', 'r', encoding='utf-8') as file:
                for line in file.readlines():
                    line = line.strip()
                    if not line:
                        continue
                    
                    if "SURREY CONCRETE - ACTIVE" in line:
                        active_section = True
                        continue
                    elif "SURREY CONCRETE - SOLD OUT" in line:
                        break
                    
                    if active_section:
                        if "Project Name" in line:
                            headers = [col.strip() for col in line.split(',')]
                            continue
                        
                        if headers and len(line.split(',')) >= len(headers):
                            cols = [col.strip() for col in line.split(',')]
                            data = dict(zip(headers, cols))
                            if data.get('Project Name'):
                                project_data[data['Project Name']] = {
                                    'total_units': int(data['Total Units']),
                                    'units_sold': int(data['Units Sold']),
                                    'standing_units': int(data['Standing units']) if data.get('Standing units') else 0,
                                    'sales_start': data['Sales Start'],
                                    'completion': data['Completion']
                                }
            
            # Get pricing data from Surrey_pricing.csv
            pricing_df = pd.read_csv('data/raw/Surrey_pricing.csv')
            
            row = 1
            
            # Add Competitive Analysis header
            ws[f'A{row}'] = 'Competitive Analysis'
            self._apply_style(ws[f'A{row}'], self.styles['header'])
            row += 2
            
            # Process pricing data by project
            project_pricing = self._process_pricing_data(pricing_df)
            
            # Process each project once
            processed_projects = set()
            for project_name, metrics in project_data.items():
                if project_name in processed_projects:
                    continue
                
                processed_projects.add(project_name)
                
                # Project header
                ws[f'A{row}'] = f"{project_name} Overview"
                self._apply_style(ws[f'A{row}'], self.styles['subheader'])
                row += 2
                
                # Calculate current absorption
                total_units = metrics['total_units']
                units_sold = metrics['units_sold']
                current_absorption = (units_sold / total_units * 100) if total_units > 0 else 0
                
                # Project metrics
                project_metrics = {
                    'Total Units': f"{total_units:,}",
                    'Current Absorption': f"{current_absorption:.1f}%",
                    'Standing Inventory': f"{metrics['standing_units']:,}",
                    'Launch': metrics['sales_start'],
                    'Completion': metrics['completion']
                }
                
                self._add_metric_group(ws, row, 'Project Metrics', project_metrics)
                row += len(project_metrics) + 3
                
                row += 3  # Space between projects
            
        except Exception as e:
            print(f"Error in _add_unit_type_metrics: {str(e)}")
            traceback.print_exc()

    def _process_pricing_data(self, pricing_df) -> dict:
        """Process pricing data into project-specific format"""
        project_pricing = {}
        current_project = None
        
        # Map project names to standardized names
        project_mapping = {
            'The Manhattan': ['manhattan', 'the manhattan'],
            'Parkway 2 - Intersect': ['parkway 2', 'parkway2', 'parkway 2 - intersect'],
            'Juno': ['juno'],
            'Sequoia': ['sequoia', 'sequoia1'],
            'Georgetown Two': ['georgetown two', 'georgetown 2'],
            'Century City Holland Park - Park Tower 1': ['century city', 'century city holland park', 'tower 1'],
            'Parkway 1 - Aspect': ['parkway 1', 'parkway1', 'aspect']
        }
        
        # First pass: group data by project
        for _, row in pricing_df.iterrows():
            if pd.notna(row['source_url']):
                if not str(row['source_url']).startswith('http'):
                    current_project = row['source_url'].strip()
                    # Initialize project data if not exists
                    for std_name, variants in project_mapping.items():
                        if any(variant in current_project.lower() for variant in variants):
                            if std_name not in project_pricing:
                                project_pricing[std_name] = []
                            current_project = std_name
                            break
                elif current_project and pd.notna(row['beds']):
                    try:
                        # Clean price data
                        price_str = str(row['price']).replace('$', '').replace(',', '') if pd.notna(row['price']) else None
                        price = float(price_str) if price_str else None
                        
                        # Clean sqft data
                        sqft = float(str(row['sqft']).replace(',', '')) if pd.notna(row['sqft']) else None
                        
                        # Clean PSF data
                        psf_str = str(row['psf']).lower().strip()
                        if psf_str not in ['unknown', 'null', 'n/a', '', 'nan']:
                            psf = float(psf_str.replace('$', '').replace(',', '').replace('"', ''))
                        elif price and sqft and sqft > 0:
                            psf = price / sqft
                        else:
                            psf = None
                        
                        if psf:  # Only add if we have valid PSF
                            project_pricing[current_project].append({
                                'beds': int(row['beds']),
                                'sqft': sqft,
                                'price': price,
                                'psf': psf
                            })
                    except (ValueError, TypeError) as e:
                        print(f"Warning: Could not process row for {current_project}: {str(e)}")
                        continue
        
        return project_pricing

    def _clean_unit_data(self, price_row) -> dict:
        """Clean and validate unit data"""
        price_str = str(price_row['price']).replace('$', '').replace(',', '') if pd.notna(price_row['price']) else None
        price = float(price_str) if price_str else None
        
        sqft = float(str(price_row['sqft']).replace(',', '')) if pd.notna(price_row['sqft']) else None
        
        psf_str = str(price_row['psf']).lower().strip()
        if psf_str not in ['unknown', 'null', 'n/a', '', 'nan']:
            psf = float(psf_str.replace('$', '').replace(',', '').replace('"', ''))
        elif price and sqft and sqft > 0:
            psf = price / sqft
        else:
            return None
        
        return {
            'beds': int(price_row['beds']),
            'sqft': sqft,
            'price': price,
            'psf': psf
        }

    def _add_unit_type_table(self, ws, row: int, project_name: str, units: List[dict]) -> int:
        """Add unit type analysis table"""
        # Add Unit Type Analysis header
        ws[f'A{row}'] = 'Unit Type Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 2
        
        # Add table headers
        headers = ['Unit Type', 'Size Range', 'Price Range', 'PSF']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self._apply_style(ws.cell(row=row, column=col), self.styles['subheader'])
        row += 1
        
        # Group pricing data by unit type
        unit_type_data = {}
        if units:  # If we have pricing data
            for unit in units:
                beds = unit['beds']
                if beds not in unit_type_data:
                    unit_type_data[beds] = []
                unit_type_data[beds].append(unit)
        
        # Process all unit types (0-3 beds)
        bed_names = {0: 'Studios', 1: 'One Bed', 2: 'Two Bed', 3: 'Three Bed'}
        
        for beds in range(4):  # 0 to 3 beds
            units = unit_type_data.get(beds, [])
            
            if units:
                # Calculate metrics from available data
                valid_sqft = [u['sqft'] for u in units if u['sqft']]
                valid_price = [u['price'] for u in units if u['price']]
                valid_psf = [u['psf'] for u in units if u['psf']]
                
                if valid_sqft and valid_price and valid_psf:
                    sqft_range = f"{min(valid_sqft):.0f}-{max(valid_sqft):.0f}"
                    price_range = f"${min(valid_price):,.0f}-${max(valid_price):,.0f}"
                    avg_psf = f"${np.mean(valid_psf):.2f}"
                else:
                    sqft_range = price_range = avg_psf = "N/A"
            else:
                sqft_range = price_range = avg_psf = "N/A"
            
            # Add row
            col = 1
            ws.cell(row=row, column=col, value=bed_names.get(beds, f"{beds} Bed")); col += 1
            ws.cell(row=row, column=col, value=sqft_range); col += 1
            ws.cell(row=row, column=col, value=price_range); col += 1
            ws.cell(row=row, column=col, value=avg_psf)
            row += 1
        
        return row + 1  # Add extra space after table

    def _add_kpi_section(self, ws) -> None:
        """Add key performance indicators section"""
        row = 3
        ws[f'A{row}'] = 'Key Performance Indicators'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2

        # Market metrics with better spacing
        market_metrics = {
            'Active Project Units': f"{self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']['active_units']:,} units",
            'Sold Out Project Units': f"{self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']['sold_units']:,} units",
            'Standing Inventory': f"{self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']['standing_units']:,} units",
            'Total Pipeline Units': f"{self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']['total_units']:,} units",
            'Current Absorption Rate': f"{self.market_analyzer.analysis_results['absorption_analysis']['current_rate']:.1f}%",
            'Average Price PSF': f"${self.market_analyzer.analysis_results['pricing_analysis']['current_metrics']['avg_psf']:,.0f}",
            'Market Score': f"{self.market_analyzer.analysis_results['market_score']:.1f}/10"
        }
        
        # Add metrics with increased column widths
        ws.column_dimensions['A'].width = 25  # Wider column for metric names
        ws.column_dimensions['B'].width = 20  # Wider column for values
        
        self._add_metric_group(ws, row, 'Market Metrics', market_metrics)
        row += len(market_metrics) + 3

        # Unit Type Performance with better spacing
        unit_analysis = self.analysis_results['unit_type_analysis']
        for unit_type in ['studios', 'one_bed', 'two_bed', 'three_bed']:
            metrics = unit_analysis[unit_type]
            unit_kpis = {
                'Absorption Rate': f"{metrics['inventory_metrics']['absorption_rate']['monthly']:.1f}% monthly",
                'Annualized Rate': f"{metrics['inventory_metrics']['absorption_rate']['annualized']:.1f}% annual",
                'Average PSF': f"${metrics['pricing_metrics']['avg_psf']:.2f}",
                'Demand Index': f"{metrics['performance_metrics']['demand_index']:.2f}"
            }
            
            # Add spacing between unit type sections
            row += 1
            self._add_metric_group(ws, row, f"{unit_type.replace('_', ' ').title()} Performance", unit_kpis)
            row += len(unit_kpis) + 2

    def _add_market_overview_charts(self, ws) -> None:
        """Add market overview charts"""
        # Add absorption trend chart
        self._add_absorption_chart(ws, "E2")
        
        # Add price trend chart
        self._add_price_trends_chart(ws, "E17")
        
        # Add supply pipeline chart
        self._add_supply_pipeline_chart(ws, "E32")

    def _create_sensitivity_analysis(self) -> None:
        """Create sensitivity analysis sheet with detailed heatmaps"""
        ws = self.wb.create_sheet("Sensitivity Analysis")
        
        # Add title and introduction
        ws['A1'] = 'Market Sensitivity Analysis'
        self._apply_style(ws['A1'], self.styles['header'])
        row = 3
        
        # Overall Market Sensitivity
        self._add_market_sensitivity_section(ws, row)
        
        # Unit Type Sensitivity
        row = ws.max_row + 5
        self._add_unit_type_sensitivity_section(ws, row)
        
        # Interest Rate Sensitivity
        row = ws.max_row + 5
        self._add_completion_impact_section(ws, row)

    def _add_market_sensitivity_section(self, ws, row: int) -> None:
        """Add market-level sensitivity analysis"""
        ws[f'A{row}'] = 'Overall Market Sensitivity'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 2
        
        # Create overall sensitivity heatmap
        sensitivity_data = self.market_analyzer.analysis_results['sensitivity_analysis']
        self._create_sensitivity_heatmaps(ws, sensitivity_data, row, "Market")
        
        # Add key findings
        row = ws.max_row + 5
        findings = {
            'Price Sensitivity': "±5% price change impacts absorption by 10-15%",
            'Interest Rate Impact': "A 1% rate decrease increases absorption by 14%",
            'Supply Impact': "10% supply increase reduces absorption by 5%",
            'Critical Thresholds': "Absorption drops significantly above $1200 PSF"
        }
        self._add_metric_group(ws, row, 'Key Findings', findings)

    def _add_unit_type_sensitivity_section(self, ws, row: int) -> None:
        """Add unit type sensitivity analysis"""
        ws[f'A{row}'] = 'Unit Type Sensitivity Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        unit_analysis = self.market_analyzer.analysis_results['unit_type_analysis']
        
        for unit_type in ['studios', 'one_bed', 'two_bed', 'three_bed']:
            metrics = unit_analysis[unit_type]
            
            # Create unit type header
            ws[f'A{row}'] = f'{unit_type.replace("_", " ").title()} Sensitivity'
            self._apply_style(ws[f'A{row}'], self.styles['subheader'])
            row += 2
            
            # Calculate and display sensitivity metrics
            sensitivity = self._calculate_unit_type_sensitivity(unit_type)
            
            # Create sensitivity heatmaps
            row = self._create_sensitivity_heatmaps(ws, sensitivity, row, unit_type)
            
            # Add unit-specific findings
            findings = self._generate_unit_sensitivity_findings(unit_type, metrics)
            self._add_metric_group(ws, row, 'Key Findings', findings)
            row += len(findings) + 5

    def _add_completion_impact_section(self, ws, row: int) -> None:
        """Add completion timeline impact analysis"""
        ws[f'A{row}'] = 'Completion Timeline Impact Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        # Map completions
        completions = {
            'The Manhattan': {'date': '2029-10-01', 'units': 422},
            'Parkway 2': {'date': '2028-12-31', 'units': 396},
            'Juno': {'date': '2028-06-30', 'units': 341},
            'Sequoia': {'date': '2027-03-31', 'units': 386},
            'Georgetown': {'date': '2026-03-01', 'units': 355},
            'Century Central': {'date': '2025-09-01', 'units': 409},
            'Parkway 1': {'date': '2026-12-31', 'units': 363}
        }
        
        # Analyze completion impacts
        yearly_impact = self._analyze_yearly_completion_impact(completions)
        
        # Add yearly impact analysis
        for year, impact in sorted(yearly_impact.items()):
            ws[f'A{row}'] = f'{year} Impact Analysis'
            self._apply_style(ws[f'A{row}'], self.styles['subheader'])
            row += 2
            
            metrics = {
                'New Supply': f"{impact['new_supply']:,} units",
                'Market Impact': impact['impact_level'],
                'Price Pressure': impact['price_pressure'],
                'Secondary Market': f"Est. {impact['secondary_units']:,} resale units"
            }
            self._add_metric_group(ws, row, 'Impact Metrics', metrics)
            row += len(metrics) + 3

    def _analyze_yearly_completion_impact(self, completions: Dict) -> Dict:
        """Analyze completion impact by year"""
        yearly_impact = {}
        
        for project, details in completions.items():
            year = details['date'][:4]
            if year not in yearly_impact:
                yearly_impact[year] = {
                    'new_supply': 0,
                    'projects': [],
                    'secondary_units': 0
                }
            
            yearly_impact[year]['new_supply'] += details['units']
            yearly_impact[year]['projects'].append(project)
            yearly_impact[year]['secondary_units'] = int(details['units'] * 0.15)  # Assume 15% investor resale
        
        # Add impact levels
        for year, impact in yearly_impact.items():
            if impact['new_supply'] > 1000:
                impact['impact_level'] = 'High'
                impact['price_pressure'] = 'Significant'
            elif impact['new_supply'] > 500:
                impact['impact_level'] = 'Medium'
                impact['price_pressure'] = 'Moderate'
            else:
                impact['impact_level'] = 'Low'
                impact['price_pressure'] = 'Limited'
        
        return yearly_impact

    def _generate_unit_sensitivity_findings(self, unit_type: str, metrics: Dict) -> Dict:
        """Generate sensitivity findings for a unit type without artificial floor"""
        try:
            # Calculate absorption sensitivity (remove 15% floor)
            absorption_rate = metrics['inventory_metrics']['absorption_rate']['monthly']
            price_premium = metrics['performance_metrics'].get('price_premium', 0)
            
            # Calculate price sensitivity based on actual market data
            price_sensitivity = abs(price_premium / absorption_rate) if absorption_rate > 0 else 0
            
            return {
                'absorption_rate': absorption_rate,
                'price_premium': price_premium,
                'price_sensitivity': price_sensitivity,
                'findings': self._generate_sensitivity_commentary(
                    unit_type,
                    metrics['performance_metrics'].get('price_premium', 0)
                )
            }
            
        except Exception as e:
            print(f"Error generating sensitivity findings: {str(e)}")
            return {
                'absorption_rate': 0,
                'price_premium': 0,
                'price_sensitivity': 0,
                'findings': []
            }

    def _get_price_sensitivity_finding(self, unit_type: str, absorption_rate: float) -> str:
        """Get price sensitivity finding for unit type"""
        if unit_type == 'studios':
            return "Highly sensitive to price changes (±17% impact)"
        elif unit_type == 'one_bed':
            return "Moderate price sensitivity (±13% impact)"
        elif unit_type == 'two_bed':
            return "Lower price sensitivity (±8% impact)"
        else:  # three_bed
            return "Least price sensitive (±6% impact)"

    def _get_competitive_position_finding(self, unit_type: str, price_premium: float) -> str:
        """Get competitive position finding"""
        if price_premium > 5:
            return "Premium position - monitor absorption"
        elif price_premium > -5:
            return "Competitive position - maintain pricing"
        else:
            return "Value position - opportunity for increases"

    def _add_risk_opportunity_summary(self, ws) -> None:
        """Add summary of key risks and opportunities"""
        row = ws.max_row + 5
        ws[f'A{row}'] = 'Key Risks and Opportunities'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        # Add risk summary
        risk_analysis = self.simulation_results['risk_analysis']
        self._add_metric_group(ws, row, 'Risk Levels', {
            k.replace('_', ' ').title(): f"{v:.1%}"
            for k, v in risk_analysis['risk_levels'].items()
        })
        
        # Add key opportunities
        row += 8
        opportunities = self._identify_opportunities()
        ws[f'A{row}'] = 'Key Opportunities'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        for opp in opportunities:
            ws[f'A{row}'] = f"• {opp}"
            row += 1

        # Use correct market average PSF
        market_avg_psf = self.market_analyzer.analysis_results['pricing_analysis']['current_metrics']['avg_psf']
        
        self._add_metric_group(ws, row, 'Market Metrics', {
            'Current Market PSF': f"${market_avg_psf:.2f}",
            # ... other metrics ...
        })

    def _identify_opportunities(self) -> List[str]:
        """Identify key market opportunities"""
        unit_analysis = self.market_analyzer.analysis_results['unit_type_analysis']
        opportunities = []
        
        # Analyze each unit type for opportunities
        for unit_type, metrics in unit_analysis.items():
            if metrics['performance_metrics']['demand_index'] > 1.2:
                opportunities.append(
                    f"High demand potential for {unit_type.replace('_', ' ')} "
                    f"(Demand Index: {metrics['performance_metrics']['demand_index']:.2f})"
                )
        
        # Add market-level opportunities
        if self.market_analyzer.analysis_results['market_score'] > 7:
            opportunities.append("Strong overall market conditions support aggressive pricing")
        
        return opportunities

    def _apply_style(self, cell, style: Dict) -> None:
        """Apply style dictionary to cell"""
        for key, value in style.items():
            setattr(cell, key, value)

    def _add_metric_group(self, ws, row: int, title: str, metrics: Dict) -> None:
        """Add a group of metrics with consistent styling and spacing"""
        # Add title
        ws[f'A{row}'] = title
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        # Add metrics with proper alignment and spacing
        for label, value in metrics.items():
            # Metric label in column A
            ws[f'A{row}'] = label
            self._apply_style(ws[f'A{row}'], self.styles['metric_label'])
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            # Metric value in column B
            ws[f'B{row}'] = value
            self._apply_style(ws[f'B{row}'], self.styles['metric_value'])
            ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            # Set row height for better spacing
            ws.row_dimensions[row].height = 20
            
            row += 1

    def _add_absorption_chart(self, ws, position: str) -> None:
        """Add absorption trend chart"""
        chart = LineChart()
        chart.title = "Absorption Trends by Unit Type"
        chart.x_axis.title = "Unit Type"
        chart.y_axis.title = "Absorption Rate (%)"
        
        # Get data from unit type analysis
        unit_analysis = self.market_analyzer.analysis_results['unit_type_analysis']
        
        # Add data
        row = ws.max_row + 2
        
        # Add headers
        headers = ["Unit Type", "Monthly Rate", "Annualized Rate"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        row += 1
        
        # Add data for each unit type
        start_row = row
        for unit_type, metrics in unit_analysis.items():
            absorption = metrics['inventory_metrics']['absorption_rate']
            ws.cell(row=row, column=1, value=unit_type.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=absorption['monthly'])
            ws.cell(row=row, column=3, value=absorption['annualized'])
            row += 1
        
        # Create data references
        data_refs = []
        for col in range(2, 4):  # Columns B and C
            data = Reference(ws, min_col=col, min_row=start_row-1, max_row=row-1)
            data_refs.append(data)
        
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=row-1)
        
        # Add data series
        for data in data_refs:
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Style the chart
        colors = ['0000FF', 'FF0000']  # Blue, Red
        for i, series in enumerate(chart.series):
            series.graphicalProperties.line.solidFill = colors[i]
            series.graphicalProperties.line.width = 20000  # 2 pt
        
        ws.add_chart(chart, position)

    def _add_price_trends_chart(self, ws, position: str) -> None:
        """Add price trends chart"""
        chart = LineChart()
        chart.title = "Price PSF by Unit Type"
        chart.x_axis.title = "Unit Type"
        chart.y_axis.title = "Price PSF ($)"
        
        # Get data from unit type analysis
        unit_analysis = self.market_analyzer.analysis_results['unit_type_analysis']
        
        # Add data
        row = ws.max_row + 2
        ws.cell(row=row, column=1, value="Unit Type")
        ws.cell(row=row, column=2, value="Price PSF")
        row += 1
        
        for unit_type, metrics in unit_analysis.items():
            ws.cell(row=row, column=1, value=unit_type.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=metrics['pricing_metrics']['avg_psf'])
            row += 1
        
        data = Reference(ws, min_col=2, min_row=row-5, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=row-5, max_row=row-1)
        
        chart.add_data(data)
        chart.set_categories(cats)
        
        ws.add_chart(chart, position)

    def _add_supply_pipeline_chart(self, ws, position: str) -> None:
        """Add supply pipeline chart"""
        chart = BarChart()
        chart.title = "Supply Pipeline"
        chart.x_axis.title = "Status"
        chart.y_axis.title = "Number of Units"
        
        # Get supply data
        supply_data = self.market_analyzer.analysis_results['supply_analysis']
        pipeline = supply_data['current_pipeline']
        
        # Add data
        row = ws.max_row + 2
        ws.cell(row=row, column=1, value="Status")
        ws.cell(row=row, column=2, value="Units")
        row += 1
        
        # Updated pipeline data using correct metrics
        pipeline_data = [
            ("Active Projects", pipeline.get('active_units', 0)),      # 2,672 units
            ("Sold Out Projects", pipeline.get('sold_units', 0)),      # 3,205 units
            ("Standing Inventory", pipeline.get('standing_units', 0)),  # 895 units
            ("Total Pipeline", pipeline.get('total_units', 0))         # 5,877 units
        ]
        
        for status, units in pipeline_data:
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=units)
            row += 1

    def _add_unit_type_charts(self, ws) -> None:
        """Add comparative charts for unit types"""
        # Absorption Rate Comparison
        chart1 = BarChart()
        chart1.title = "Absorption Rate by Unit Type"
        chart1.y_axis.title = "Absorption Rate (%)"
        
        # Price PSF Comparison
        chart2 = BarChart()
        chart2.title = "Average PSF by Unit Type"
        chart2.y_axis.title = "Price PSF ($)"
        
        # Size Range Comparison
        chart3 = BarChart()
        chart3.title = "Size Range by Unit Type"
        chart3.y_axis.title = "Square Footage"
        
        # Add data and charts
        unit_analysis = self.market_analyzer.analysis_results['unit_type_analysis']
        categories = []
        monthly_absorption = []
        annualized_absorption = []
        psf_rates = []
        avg_sizes = []
        
        row = ws.max_row + 2
        # Add headers
        headers = ['Unit Type', 'Monthly Absorption', 'Annualized Rate', 'PSF', 'Avg Size']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self._apply_style(ws.cell(row=row, column=col), self.styles['subheader'])
        
        row += 1
        for unit_type, metrics in unit_analysis.items():
            categories.append(unit_type.replace('_', ' ').title())
            absorption = metrics['inventory_metrics']['absorption_rate']
            monthly_absorption.append(absorption['monthly'])
            annualized_absorption.append(absorption['annualized'])
            psf_rates.append(metrics['pricing_metrics']['avg_psf'])
            avg_sizes.append(np.mean(metrics['pricing_metrics']['size_range']))
            
            # Add data to worksheet
            ws.cell(row=row, column=1, value=categories[-1])
            ws.cell(row=row, column=2, value=monthly_absorption[-1])
            ws.cell(row=row, column=3, value=annualized_absorption[-1])
            ws.cell(row=row, column=4, value=psf_rates[-1])
            ws.cell(row=row, column=5, value=avg_sizes[-1])
            row += 1
        
        # Create data references for charts
        data_monthly = Reference(ws, min_col=2, min_row=row-5, max_row=row-1)
        data_annual = Reference(ws, min_col=3, min_row=row-5, max_row=row-1)
        data_psf = Reference(ws, min_col=4, min_row=row-5, max_row=row-1)
        data_size = Reference(ws, min_col=5, min_row=row-5, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=row-5, max_row=row-1)
        
        # Add data to charts
        chart1.add_data(data_monthly, titles_from_data=True)
        chart1.add_data(data_annual, titles_from_data=True)
        chart1.set_categories(cats)
        ws.add_chart(chart1, "E2")
        
        chart2.add_data(data_psf)
        chart2.set_categories(cats)
        ws.add_chart(chart2, "E17")
        
        chart3.add_data(data_size)
        chart3.set_categories(cats)
        ws.add_chart(chart3, "E32")

    def _add_unit_type_summary(self, ws) -> None:
        """Add unit type summary section to executive summary"""
        row = ws.max_row + 5
        ws[f'A{row}'] = 'Unit Type Performance Summary'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        unit_analysis = self.market_analyzer.analysis_results['unit_type_analysis']
        
        # Add summary table headers
        headers = ['Unit Type', 'Monthly Absorption', 'Annualized Rate', 'Avg PSF', 'Demand Index']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['subheader'])
        row += 1
        
        # Add data for each unit type
        for unit_type in ['studios', 'one_bed', 'two_bed', 'three_bed']:
            metrics = unit_analysis[unit_type]
            absorption = metrics['inventory_metrics']['absorption_rate']
            
            # Unit type name
            ws.cell(row=row, column=1, value=unit_type.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=f"{absorption['monthly']:.1f}")
            self._apply_absorption_formatting(ws.cell(row=row, column=2), absorption['monthly'])
            
            # Annualized absorption rate
            annual_cell = ws.cell(row=row, column=3, value=f"{absorption['annualized']:.1f}%")
            self._apply_absorption_formatting(annual_cell, absorption['annualized']/12)
            
            # Average PSF
            ws.cell(row=row, column=4, value=f"${metrics['pricing_metrics']['avg_psf']:.2f}")
            
            # Demand index with conditional formatting
            demand_cell = ws.cell(row=row, column=5, value=f"{metrics['performance_metrics']['demand_index']:.2f}")
            self._apply_demand_formatting(demand_cell, 
                                        metrics['performance_metrics']['demand_index'])
            
            row += 1
        
        # Add insights section
        row += 2
        ws[f'A{row}'] = 'Key Insights'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        insights = self._generate_unit_type_insights(unit_analysis)
        for insight in insights:
            ws[f'A{row}'] = f"• {insight}"
            row += 1

    def _apply_absorption_formatting(self, cell, absorption_rate: float) -> None:
        """Apply conditional formatting based on absorption rate"""
        if absorption_rate >= 70:
            cell.fill = self.styles['risk_low']
        elif absorption_rate >= 50:
            cell.fill = self.styles['risk_medium']
        else:
            cell.fill = self.styles['risk_high']

    def _apply_demand_formatting(self, cell, demand_index: float) -> None:
        """Apply conditional formatting based on demand index"""
        if demand_index >= 1.2:
            cell.fill = self.styles['risk_low']
        elif demand_index >= 0.8:
            cell.fill = self.styles['risk_medium']
        else:
            cell.fill = self.styles['risk_high']

    def _generate_unit_type_insights(self, unit_analysis: Dict) -> List[str]:
        """Generate insights from unit type analysis"""
        insights = []
        
        # Best performing unit type by absorption rate
        best_absorption = max(
            unit_analysis.items(),
            key=lambda x: x[1]['inventory_metrics']['absorption_rate']['monthly']
        )
        insights.append(
            f"{best_absorption[0].replace('_', ' ').title()} shows strongest absorption "
            f"at {best_absorption[1]['inventory_metrics']['absorption_rate']['monthly']:.1f}% monthly "
            f"({best_absorption[1]['inventory_metrics']['absorption_rate']['annualized']:.1f}% annualized)"
        )
        
        # Price premium analysis
        highest_premium = max(
            unit_analysis.items(),
            key=lambda x: x[1]['performance_metrics']['price_premium']
        )
        insights.append(
            f"{highest_premium[0].replace('_', ' ').title()} commands highest price premium "
            f"at {highest_premium[1]['performance_metrics']['price_premium']:.1f}%"
        )
        
        # Demand opportunities
        high_demand_types = [
            unit_type for unit_type, metrics in unit_analysis.items()
            if metrics['performance_metrics']['demand_index'] > 1.2
        ]
        if high_demand_types:
            types_list = ', '.join(t.replace('_', ' ').title() for t in high_demand_types)
            insights.append(f"High demand potential for {types_list}")
        
        # Supply insights
        total_available = sum(
            metrics['inventory_metrics']['available_units'] 
            for metrics in unit_analysis.values()
        )
        low_inventory_types = [
            unit_type for unit_type, metrics in unit_analysis.items()
            if metrics['inventory_metrics']['available_units'] < total_available * 0.15
        ]
        if low_inventory_types:
            types_list = ', '.join(t.replace('_', ' ').title() for t in low_inventory_types)
            insights.append(f"Limited inventory available for {types_list}")
        
        return insights

    def _add_competitive_analysis(self, ws) -> None:
        """Add detailed competitive analysis section"""
        try:
            # Get project data from Surrey_Concrete_Launches_correct.csv
            project_data = {}
            active_section = False
            headers = None
            
            with open('data/raw/Surrey_Concrete_Launches_correct.csv', 'r', encoding='utf-8') as file:
                for line in file.readlines():
                    line = line.strip()
                    if not line:
                        continue
                    
                    if "SURREY CONCRETE - ACTIVE" in line:
                        active_section = True
                        continue
                    elif "SURREY CONCRETE - SOLD OUT" in line:
                        break
                    
                    if active_section:
                        if "Project Name" in line:
                            headers = [col.strip() for col in line.split(',')]
                            continue
                        
                        if headers and len(line.split(',')) >= len(headers):
                            cols = [col.strip() for col in line.split(',')]
                            data = dict(zip(headers, cols))
                            if data.get('Project Name'):
                                project_data[data['Project Name']] = {
                                    'total_units': int(data['Total Units']),
                                    'units_sold': int(data['Units Sold']),
                                    'standing_units': int(data['Standing units']) if data.get('Standing units') else 0,
                                    'sales_start': data['Sales Start'],
                                    'completion': data['Completion']
                                }
        
            row = 1
            
            # Add Competitive Analysis header
            ws[f'A{row}'] = 'Competitive Analysis'
            self._apply_style(ws[f'A{row}'], self.styles['header'])
            row += 2
            
            # Process each project from Surrey_Concrete_Launches_correct.csv
            for project_name, metrics in project_data.items():
                # Project header
                ws[f'A{row}'] = f"{project_name} Overview"
                self._apply_style(ws[f'A{row}'], self.styles['subheader'])
                row += 2
                
                # Calculate current absorption
                total_units = metrics['total_units']
                units_sold = metrics['units_sold']
                current_absorption = (units_sold / total_units * 100) if total_units > 0 else 0
                
                # Project metrics
                project_metrics = {
                    'Total Units': f"{total_units:,}",
                    'Current AI': f"{current_absorption:.1f}%",
                    'Standing': f"{metrics['standing_units']:,}",
                    'Launch': metrics['sales_start'],
                    'Completion': metrics['completion']
                }
                
                self._add_metric_group(ws, row, 'Project Metrics', project_metrics)
                row += len(project_metrics) + 3
        
        except Exception as e:
            print(f"Error in _add_competitive_analysis: {str(e)}")
            traceback.print_exc()

    def _get_project_metrics(self, project_name: str) -> Dict:
        """Get project metrics from market data"""
        try:
            # Find project in active projects
            project = next((p for p in self.market_analyzer.project_data['active_projects']['projects'] 
                           if p['name'] == project_name), None)
            
            if project:
                return {
                    'total_units': project.get('total_units', 0),
                    'current_absorption': (project.get('units_sold', 0) / project.get('total_units', 1)) * 100,
                    'sales_start': project.get('sales_start', 'N/A'),
                    'completion': project.get('completion', 'N/A')
                }
            
            # Fallback values if project not found
            return {
                'total_units': 0,
                'current_absorption': 0.0,
                'sales_start': 'N/A',
                'completion': 'N/A'
            }
            
        except Exception as e:
            print(f"Error getting project metrics for {project_name}: {str(e)}")
            return {
                'total_units': 0,
                'current_absorption': 0.0,
                'sales_start': 'N/A',
                'completion': 'N/A'
            }

    def _add_competitive_positioning_matrix(self, ws, row: int) -> int:
        """Add competitive positioning matrix"""
        ws[f'A{row}'] = 'Competitive Positioning Matrix'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2

        # Create positioning matrix
        matrix_data = {
            'Price Position': {
                'Manhattan': 'Premium',
                'Parkway 2': 'Market',
                'Our Position': 'Competitive Premium'
            },
            'Unit Mix': {
                'Manhattan': 'Diverse (Studio-3Bed)',
                'Parkway 2': 'Limited (1-2Bed)',
                'Our Position': 'Full Range'
            },
            'Target Market': {
                'Manhattan': 'Luxury/Investors',
                'Parkway 2': 'End Users',
                'Our Position': 'Balanced Mix'
            },
            'Timing': {
                'Manhattan': 'Oct 2024',
                'Parkway 2': 'Apr 2024',
                'Our Position': 'Strategic Launch'
            }
        }

        # Add matrix headers
        headers = ['Attribute', 'Manhattan', 'Parkway 2', 'Our Position']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['subheader'])
        row += 1

        # Add matrix data
        for attribute, positions in matrix_data.items():
            ws.cell(row=row, column=1, value=attribute)
            for col, (comp, position) in enumerate(positions.items(), 2):
                ws.cell(row=row, column=col, value=position)
            row += 1

        return row

    def _add_strategic_implications(self, ws, row: int, competitors: Dict) -> int:
        """Add strategic implications section"""
        ws[f'A{row}'] = 'Strategic Implications'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2

        # Timing implications
        timing_implications = {
            'Launch Window': 'Strategic launch between Manhattan and Parkway 2',
            'Market Entry': 'Capitalize on spring market momentum',
            'Absorption Target': '65% in first 12 months achievable with proper positioning'
        }
        self._add_metric_group(ws, row, 'Timing Strategy', timing_implications)
        row += 6

        # Pricing implications
        pricing_implications = {
            'Studios': 'Slight discount to Manhattan (-2-3%) to capture first-time buyers',
            'One Beds': 'At market with premium features for strong absorption',
            'Two Beds': 'Premium positioning (+2-3%) given limited competition',
            'Three Beds': 'Significant premium (+4-5%) for luxury segment'
        }
        self._add_metric_group(ws, row, 'Pricing Strategy', pricing_implications)
        row += 7

        # Georgetown 2 implications
        georgetown_implications = {
            'Market Impact': 'Monitor launch timing and adjust strategy accordingly',
            'Opportunity': 'Establish market position before potential launch',
            'Risk Mitigation': 'Build strong absorption momentum in initial phases'
        }
        self._add_metric_group(ws, row, 'Georgetown 2 Considerations', georgetown_implications)

        return row

    def _add_absorption_targets(self, ws, row: int) -> int:
        """Add detailed absorption targets"""
        ws[f'A{row}'] = 'Absorption Targets and Milestones'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2

        # Phase 1: Months 1-3
        phase1_targets = {
            'Target': '25-30% absorption',
            'Strategy': 'Competitive entry pricing with early buyer incentives',
            'Focus': 'High-demand 1-bed and 2-bed units',
            'Pricing': 'Initial pricing at market with selective premiums'
        }
        self._add_metric_group(ws, row, 'Phase 1 (Months 1-3)', phase1_targets)
        row += 7

        # Phase 2: Months 4-6
        phase2_targets = {
            'Target': '40-45% absorption',
            'Strategy': 'Selective price increases on high-performing units',
            'Focus': 'Premium units and remaining 1-beds',
            'Pricing': '2-3% increase on select inventory'
        }
        self._add_metric_group(ws, row, 'Phase 2 (Months 4-6)', phase2_targets)
        row += 7

        # Phase 3: Months 7-12
        phase3_targets = {
            'Target': '65% absorption',
            'Strategy': 'Value-based pricing with market adjustments',
            'Focus': 'Balanced mix across unit types',
            'Pricing': '3-4% increase on remaining inventory'
        }
        self._add_metric_group(ws, row, 'Phase 3 (Months 7-12)', phase3_targets)
        row += 7

        # Long-term strategy
        longterm_strategy = {
            'Target': 'Remaining 35% absorption',
            'Timeline': 'Months 13-48',
            'Strategy': 'Premium pricing with targeted releases',
            'Opportunity': 'Capitalize on market strengthening periods'
        }
        self._add_metric_group(ws, row, 'Long-term Strategy', longterm_strategy)

        return row

    def _calculate_price_tower(self, base_psf: float) -> Dict:
        """Calculate price tower based on market principles"""
        price_tower = {
            'studios': {
                'psf_premium': -0.05,  # 5% discount to base
                'target_sizes': [300, 350],
                'price_range': []
            },
            'one_bed': {
                'psf_premium': 0.00,   # At base
                'target_sizes': [450, 550],
                'price_range': []
            },
            'two_bed': {
                'psf_premium': 0.03,   # 3% premium to base
                'target_sizes': [700, 850],
                'price_range': []
            },
            'three_bed': {
                'psf_premium': 0.05,   # 5% premium to base
                'target_sizes': [950, 1100],
                'price_range': []
            }
        }
        
        # Calculate price ranges for each unit type
        for unit_type, specs in price_tower.items():
            adjusted_psf = base_psf * (1 + specs['psf_premium'])
            specs['price_range'] = [
                round(size * adjusted_psf / 5000) * 5000  # Round to nearest $5k
                for size in specs['target_sizes']
            ]
            specs['avg_psf'] = adjusted_psf
        
        return price_tower

    def _add_pricing_recommendations(self, ws) -> None:
        """Add pricing recommendations section"""
        row = ws.max_row + 5
        ws[f'A{row}'] = 'Pricing Recommendations'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        # Get correct market PSF from project data
        market_psf = self.market_analyzer.analysis_results['pricing_analysis']['current_metrics']['avg_psf']
        
        # Overall Market Strategy
        market_metrics = {
            'Base Market PSF': f"${market_psf:.2f}",
            'Market Position': self._determine_market_position(),
            'Price Trend': self._determine_price_trend(),
            'Launch Strategy': "Strategic April 2024 launch between competitors"
        }
        self._add_metric_group(ws, row, 'Market Pricing Strategy', market_metrics)
        row += len(market_metrics) + 3
        
        # Unit Type Specific Recommendations
        for unit_type in self.market_analyzer.unit_types:
            price_points = self.market_analyzer._calculate_unit_price_points(unit_type, self.market_analyzer.analysis_results['pricing_analysis'])
            
            ws[f'A{row}'] = f'{unit_type.replace("_", " ").title()} Pricing Strategy'
            self._apply_style(ws[f'A{row}'], self.styles['subheader'])
            row += 2
            
            recommendations = {
                'Target PSF': f"${price_points['target_psf']:.2f}",
                'Price Range': f"${price_points['min_price']:,.0f} - ${price_points['max_price']:,.0f}",
                'Premium/Discount': f"{price_points['premium']:+.1f}%",
                'Strategy': price_points['strategy'],
                'Competitive Position': self._get_competitive_position(unit_type, price_points['target_psf'], market_psf)
            }
            self._add_metric_group(ws, row, 'Recommendations', recommendations)
            row += len(recommendations) + 4

    def _get_competitive_position(self, unit_type: str, target_psf: float, competitor_psf: float) -> str:
        """Get competitive position description"""
        diff_pct = ((target_psf / competitor_psf) - 1) * 100 if competitor_psf > 0 else 0
        
        if unit_type == 'studios':
            return f"Value position at {diff_pct:.1f}% vs competitors to drive absorption"
        elif unit_type == 'one_bed':
            return f"Market position with selective premiums ({diff_pct:+.1f}%)"
        elif unit_type == 'two_bed':
            return f"Premium position targeting end-users ({diff_pct:+.1f}%)"
        else:  # three_bed
            return f"Luxury position with significant premium ({diff_pct:+.1f}%)"

    def _add_unit_type_recommendations(self, ws) -> None:
        """Add recommendations for each unit type"""
        try:
            row = ws.max_row + 5
            ws[f'A{row}'] = 'Unit Type Recommendations'
            self._apply_style(ws[f'A{row}'], self.styles['header'])
            row += 2

            # Get recommendations from market analyzer
            recommendations = self.market_analyzer._generate_unit_recommendations()
            
            if not recommendations:
                print("Warning: No unit type recommendations generated")
                return

            for unit_type in ['studios', 'one_bed', 'two_bed', 'three_bed']:
                if unit_type not in recommendations:
                    continue

                # Unit Type Header
                ws[f'A{row}'] = f'{unit_type.replace("_", " ").title()} Recommendations'
                self._apply_style(ws[f'A{row}'], self.styles['subheader'])
                row += 2

                unit_recs = recommendations[unit_type]

                # Pricing Strategy Section
                pricing_data = {
                    'Base Strategy': unit_recs['pricing_strategy']['base_strategy'],
                    'Target Premium': unit_recs['pricing_strategy']['target_premium'],
                    'Absorption Target': unit_recs['pricing_strategy']['absorption_target'],
                    'Current Absorption': unit_recs['pricing_strategy']['current_absorption']
                }
                row = self._add_recommendation_section(ws, row, 'Pricing Strategy', pricing_data)

                # Market Metrics Section
                market_data = {
                    'Market PSF': unit_recs['market_metrics']['market_psf'],
                    'Demand Index': unit_recs['market_metrics']['demand_index'],
                    'Sales Velocity': unit_recs['market_metrics']['velocity_metric'],
                    'Market Share': unit_recs['market_metrics']['market_share']
                }
                row = self._add_recommendation_section(ws, row, 'Market Metrics', market_data)

                # Revenue Analysis Section
                revenue_data = {
                    'Price Band': unit_recs['revenue_analysis']['price_band'],
                    'Revenue Contribution': unit_recs['revenue_analysis']['revenue_contribution'],
                    'Monthly Revenue': unit_recs['revenue_analysis']['absorption_revenue'],
                    'Market Depth': unit_recs['revenue_analysis']['market_depth']
                }
                row = self._add_recommendation_section(ws, row, 'Revenue Analysis', revenue_data)

                row += 2  # Add space between unit types

        except Exception as e:
            print(f"Error adding unit type recommendations: {str(e)}")
            traceback.print_exc()

    def _add_recommendation_section(self, ws, row: int, section_title: str, data: Dict) -> int:
        """Add a recommendation section with consistent formatting"""
        # Add section title
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = section_title
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1

        # Add data rows
        for key, value in data.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            
            # Apply styles
            self._apply_style(ws[f'A{row}'], self.styles['metric_label'])
            self._apply_style(ws[f'B{row}'], self.styles['metric_value'])
            
            row += 1
        
        row += 1  # Add space after section
        return row

    def _generate_unit_pricing_recommendations(self, metrics: Dict, unit_type: str) -> Dict:
        """Generate pricing recommendations for specific unit type"""
        absorption = metrics['inventory_metrics']['absorption_rate']
        monthly_absorption = absorption['monthly']  # Get monthly absorption rate
        price_premium = metrics['performance_metrics']['price_premium']
        demand_index = metrics['performance_metrics']['demand_index']
        
        # Base recommendations on unit type and metrics
        if unit_type == 'studios':
            base_strategy = "Entry-level pricing to capture first-time buyers"
            target_premium = -2.0 if monthly_absorption < 50 else 0.0
        elif unit_type == 'one_bed':
            base_strategy = "Market-rate pricing with selective premiums"
            target_premium = 0.0 if monthly_absorption < 60 else 2.0
        elif unit_type == 'two_bed':
            base_strategy = "Premium positioning for end-users"
            target_premium = 2.0 if monthly_absorption < 70 else 3.0
        else:  # three_bed
            base_strategy = "Luxury positioning for premium buyers"
            target_premium = 3.0 if monthly_absorption < 65 else 5.0
        
        return {
            'Base Strategy': base_strategy,
            'Target Premium': f"{target_premium:+.1f}% to market",
            'Absorption Target': f"{min(monthly_absorption + 10, 75):.1f}%",
            'Price Adjustment': self._get_price_adjustment_recommendation(monthly_absorption, price_premium)
        }

    def _generate_unit_product_recommendations(self, metrics: Dict, unit_type: str) -> Dict:
        """Generate product recommendations for specific unit type"""
        size_range = metrics['pricing_metrics']['size_range']
        avg_size = np.mean(size_range)
        
        # Base recommendations on unit type
        if unit_type == 'studios':
            layout_focus = "Efficient layouts with multi-functional spaces"
            target_features = "Smart home features, built-in storage"
        elif unit_type == 'one_bed':
            layout_focus = "Open concept with flexible spaces"
            target_features = "Home office nook, premium appliances"
        elif unit_type == 'two_bed':
            layout_focus = "Family-oriented layouts with separation"
            target_features = "Large balconies, upgraded finishes"
        else:  # three_bed
            layout_focus = "Luxury layouts with entertainment spaces"
            target_features = "Premium upgrades, expansive views"
        
        return {
            'Layout Focus': layout_focus,
            'Target Features': target_features,
            'Size Strategy': self._get_size_recommendation(avg_size, unit_type),
            'View Premium': self._get_view_premium_recommendation(unit_type)
        }

    def _generate_unit_strategy_recommendations(self, metrics: Dict, unit_type: str) -> Dict:
        """Generate sales strategy recommendations for specific unit type"""
        absorption = metrics['inventory_metrics']['absorption_rate']
        monthly_absorption = absorption['monthly']  # Get monthly absorption rate
        available = metrics['inventory_metrics']['available_units']
        
        # Base strategy on unit type and performance
        if monthly_absorption < 40:
            release_strategy = "Conservative releases with incentives"
            pricing_approach = "Competitive entry pricing"
        elif monthly_absorption < 60:
            release_strategy = "Measured releases based on demand"
            pricing_approach = "Market-based pricing with selective premiums"
        else:
            release_strategy = "Accelerated releases to maintain momentum"
            pricing_approach = "Premium pricing with regular increases"
        
        return {
            'Release Strategy': release_strategy,
            'Pricing Approach': pricing_approach,
            'Inventory Management': self._get_inventory_recommendation(available, monthly_absorption),
            'Target Market': self._get_target_market_recommendation(unit_type)
        }

    def _get_price_adjustment_recommendation(self, absorption_rate: float, price_premium: float) -> str:
        """Get price adjustment recommendation based on metrics"""
        if absorption_rate < 40 and price_premium > 0:
            return "Consider price adjustments or incentives"
        elif absorption_rate < 60 and price_premium > 5:
            return "Monitor market response to premium"
        elif absorption_rate >= 70:
            return "Opportunity for selective increases"
        else:
            return "Maintain current pricing strategy"

    def _get_size_recommendation(self, avg_size: float, unit_type: str) -> str:
        """Get size recommendation based on unit type"""
        typical_sizes = {
            'studios': 450,
            'one_bed': 600,
            'two_bed': 850,
            'three_bed': 1100
        }
        
        diff_pct = ((avg_size - typical_sizes[unit_type]) / typical_sizes[unit_type]) * 100
        
        if diff_pct < -10:
            return "Consider increasing average size"
        elif diff_pct < -5:
            return "Slightly below market average"
        elif diff_pct < 5:
            return "Competitive size offering"
        else:
            return "Premium size positioning"

    def _get_view_premium_recommendation(self, unit_type: str) -> str:
        """Get view premium recommendation based on unit type"""
        if unit_type in ['three_bed', 'two_bed']:
            return "5-8% for optimal views"
        elif unit_type == 'one_bed':
            return "3-5% for preferred views"
        else:
            return "2-3% for select views"

    def _get_inventory_recommendation(self, available: int, absorption: float) -> str:
        """Get inventory management recommendation"""
        if available < 5:
            return "Priority for new releases"
        elif available < 10:
            return "Monitor inventory levels"
        elif absorption < 50:
            return "Consider phased releases"
        else:
            return "Maintain current strategy"

    def _get_target_market_recommendation(self, unit_type: str) -> str:
        """Get target market recommendation based on unit type"""
        targets = {
            'studios': "First-time buyers and investors",
            'one_bed': "Young professionals and downsizers",
            'two_bed': "End-users and young families",
            'three_bed': "Move-up buyers and luxury segment"
        }
        return targets.get(unit_type, "Mixed demographic")

    def _create_market_analysis(self) -> None:
        """Create market analysis sheet"""
        ws = self.wb.create_sheet("Market Analysis")
        
        # Market Overview
        self._add_market_overview(ws)
        
        # Supply Analysis
        self._add_supply_analysis(ws)
        
        # Price Analysis
        self._add_price_analysis(ws)
        
        # Absorption Analysis
        self._add_absorption_analysis(ws)
        
        # Market Factors Analysis
        self._add_market_factors_analysis(ws)

    def _add_market_overview(self, ws) -> None:
        """Add market overview section"""
        ws['A1'] = 'Market Overview'
        self._apply_style(ws['A1'], self.styles['header'])
        
        row = 3
        # Use the same data source as KPI section
        supply_data = self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']
        market_metrics = {
            'Active Project Units': f"{supply_data['active_units']:,} units",  # 2,672 units
            'Sold Out Project Units': f"{supply_data['sold_units']:,} units",  # 3,205 units
            'Standing Inventory': f"{supply_data['standing_units']:,} units",  # 895 units
            'Total Pipeline Units': f"{supply_data['total_units']:,} units"    # 5,877 units
        }
        
        self._add_metric_group(ws, row, 'Key Metrics', market_metrics)

    def _add_supply_analysis(self, ws) -> None:
        """Add supply analysis section with quarterly distribution"""
        row = ws.max_row + 5
        ws[f'A{row}'] = 'Supply Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        # Use the same data source consistently
        supply_data = self.market_analyzer.analysis_results['supply_analysis']
        pipeline = supply_data['current_pipeline']
        
        # Current pipeline
        pipeline_metrics = {
            'Active Project Units': f"{pipeline['active_units']:,} units",     # 2,672 units
            'Sold Out Project Units': f"{pipeline['sold_units']:,} units",     # 3,205 units
            'Standing Inventory': f"{pipeline['standing_units']:,} units",     # 895 units
            'Total Units': f"{pipeline['total_units']:,} units"       # 5,877 units
        }
        self._add_metric_group(ws, row, 'Current Pipeline', pipeline_metrics)
        row += len(pipeline_metrics) + 3
        
        # Add quarterly supply distribution
        ws[f'A{row}'] = 'Quarterly Supply Distribution (by Sales Start)'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 2
        
        # Add headers
        headers = ['Quarter', 'Total Units', 'Standing Units', 'Projects', 'Status']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self._apply_style(ws.cell(row=row, column=col), self.styles['subheader'])
        row += 1
        
        # Add quarterly data
        quarterly_supply = supply_data.get('quarterly_distribution', {})
        for quarter in sorted(quarterly_supply.keys()):
            data = quarterly_supply[quarter]
            projects_str = ", ".join(f"{p['name']} ({p['units']} units)" for p in data['projects'])
            
            col = 1
            ws.cell(row=row, column=col, value=quarter); col += 1
            ws.cell(row=row, column=col, value=data['total_units']); col += 1
            ws.cell(row=row, column=col, value=data['standing_units']); col += 1
            ws.cell(row=row, column=col, value=projects_str); col += 1
            ws.cell(row=row, column=col, value=data['status'])
            
            # Apply alternating row colors
            for cell in ws[row]:
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            row += 1
        
        # Add total row
        row += 1
        ws.cell(row=row, column=1, value='Total')
        ws.cell(row=row, column=2, value=sum(q['total_units'] for q in quarterly_supply.values()))
        ws.cell(row=row, column=3, value=sum(q['standing_units'] for q in quarterly_supply.values()))
        self._apply_style(ws.cell(row=row, column=1), self.styles['subheader'])
        self._apply_style(ws.cell(row=row, column=2), self.styles['subheader'])
        self._apply_style(ws.cell(row=row, column=3), self.styles['subheader'])

    def _add_price_analysis(self, ws) -> None:
        """Add price analysis section"""
        row = ws.max_row + 5
        ws[f'A{row}'] = 'Price Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        pricing_data = self.market_analyzer.analysis_results['pricing_analysis']
        current_metrics = pricing_data['current_metrics']
        
        # Current metrics
        price_metrics = {
            'Average PSF': f"${current_metrics['avg_psf']:.2f}",
            'Price Range': f"${current_metrics['price_range']['min']:,.0f} - ${current_metrics['price_range']['max']:,.0f}",
            'Median Price': f"${current_metrics['price_range']['median']:,.0f}"
        }
        self._add_metric_group(ws, row, 'Current Pricing', price_metrics)
        row += 5
        
        # Project pricing comparison for all competitors
        ws[f'A{row}'] = 'Project Pricing Comparison'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 2
        
        headers = ['Project', 'PSF', 'Premium/Discount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['subheader'])
        row += 1
        
        # Get all competitor data
        competitor_data = {}
        for unit_type in ['studios', 'one_bed', 'two_bed', 'three_bed']:
            comp_data = self.market_analyzer._get_competitor_performance(unit_type)
            for comp in comp_data:
                if comp['name'] not in competitor_data and comp['name'] != 'Unknown':
                    competitor_data[comp['name']] = comp
        
        market_avg = current_metrics['avg_psf']
        for project_name, data in competitor_data.items():
            project_psf = data['psf']
            premium = ((project_psf / market_avg) - 1) * 100 if market_avg > 0 else 0
            
            ws.cell(row=row, column=1, value=project_name)
            ws.cell(row=row, column=2, value=f"${project_psf:.2f}")
            ws.cell(row=row, column=3, value=f"{premium:+.1f}%")
            row += 1

    def _add_absorption_analysis(self, ws) -> None:
        """Add absorption analysis section"""
        row = ws.max_row + 5
        ws[f'A{row}'] = 'Absorption Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2

        # Market Performance (Current)
        market_metrics = {
            'Market Monthly Rate': f"{self.market_analyzer.analysis_results['absorption_analysis']['market_average']['monthly_rate']:.1f}%",
            'Market Annualized': f"{self.market_analyzer.analysis_results['absorption_analysis']['market_average']['annualized_rate']:.1f}%",
            'Target Monthly Rate': f"{self.market_analyzer.analysis_results['absorption_analysis']['target']['base_monthly']:.1f}%",
            'Target Annual Rate': f"{self.market_analyzer.analysis_results['absorption_analysis']['target']['annual_target']:.1f}%"
        }
        self._add_metric_group(ws, row, 'Market Performance', market_metrics)
        row += 6

        # Competitor Performance
        competitor_metrics = {
            'Manhattan Monthly Rate': f"{self.market_analyzer.analysis_results['absorption_analysis']['competitor_performance']['manhattan']['monthly_rate']:.1f}%",
            'Manhattan Total': f"{self.market_analyzer.analysis_results['absorption_analysis']['competitor_performance']['manhattan']['total_absorption']:.1f}%",
            'Parkway 2 Monthly Rate': f"{self.market_analyzer.analysis_results['absorption_analysis']['competitor_performance']['parkway2']['monthly_rate']:.1f}%",
            'Parkway 2 Total': f"{self.market_analyzer.analysis_results['absorption_analysis']['competitor_performance']['parkway2']['total_absorption']:.1f}%"
        }
        self._add_metric_group(ws, row, 'Competitor Performance', competitor_metrics)
        row += 6

        # Project Absorption (Future)
        ws[f'A{row}'] = 'Note: Project absorption metrics will be available after launch in April 2025'
        self._apply_style(ws[f'A{row}'], self.styles['metric_label'])

    def _add_market_factors_analysis(self, ws) -> None:
        """Add market factors analysis"""
        row = ws.max_row + 5
        ws[f'A{row}'] = 'Market Factors Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        market_factors = self.market_analyzer.analysis_results['market_factors']
        
        # Interest Rates
        interest_rates = market_factors['interest_rates']
        current_rate = interest_rates['current']['rates']['5yr_fixed']  # Access the nested structure correctly
        historical = interest_rates['historical_trends']['5yr_fixed']
        
        # Calculate trend from historical data
        rate_trend = "Decreasing" if historical['2024_avg'] < historical['2023_avg'] else "Increasing"
        
        self._add_metric_group(ws, row, 'Interest Rates', {
            'Current 5yr Fixed': f"{current_rate:.2f}%",
            'YTD Average': f"{historical['2024_avg']:.2f}%",
            'Previous Year': f"{historical['2023_avg']:.2f}%",
            'Rate Trend': rate_trend
        })
        row += 5
        
        # Employment
        employment = market_factors['employment']
        current_emp = employment['current_statistics']
        emp_trends = employment['historical_trends']['2024_ytd']['employment_rate']
        
        self._add_metric_group(ws, row, 'Employment', {
            'Employment Rate': f"{current_emp['employment_rate']:.1%}",
            'Unemployment Rate': f"{current_emp['unemployment_rate']:.1%}",
            'YTD Average': f"{emp_trends['average']:.1%}",
            'Trend': emp_trends['trend']
        })
        row += 5
        
        # Demographics
        demographics = market_factors['demographics']['household_income']
        current_income = demographics['current']
        growth = demographics['growth_metrics']['average_income_growth']
        
        self._add_metric_group(ws, row, 'Demographics', {
            'Average Income': f"${current_income['average_income']:,.0f}",
            'Median Income': f"${current_income['median_income']:,.0f}",
            'Income Growth': f"{growth['percentage']:.1f}%",
            'Annual Growth Rate': f"{growth['cagr']:.1f}%"
        })

    def _create_risk_analysis(self) -> None:
        """Create risk analysis sheet using data from market analyzer"""
        ws = self.wb.create_sheet("Risk Analysis")
        row = 1
        
        # Get risk analysis from market analyzer
        risk_analysis = self.market_analyzer._analyze_risks()
        
        # Risk Overview
        ws[f'A{row}'] = 'Risk Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        # Add each risk category
        for risk_type in ['absorption_risk', 'price_risk', 'supply_risk', 'market_risk']:
            if risk_type in risk_analysis:
                risk_data = risk_analysis[risk_type]
                
                # Add risk type header
                ws[f'A{row}'] = risk_type.replace('_', ' ').title()
                self._apply_style(ws[f'A{row}'], self.styles['subheader'])
                row += 1
                
                # Add assessment
                if 'assessment' in risk_data:
                    ws[f'A{row}'] = 'Assessment'
                    ws[f'B{row}'] = risk_data['assessment']
                    row += 2
                
                # Add impact analysis
                if 'impact_analysis' in risk_data:
                    ws[f'A{row}'] = 'Impact Analysis'
                    row += 1
                    for point in risk_data['impact_analysis']:
                        ws[f'B{row}'] = f"• {point}"
                        row += 1
                    row += 1
                
                # Add risk factors
                if 'risk_factors' in risk_data:
                    ws[f'A{row}'] = 'Risk Factors'
                    row += 1
                    for factor in risk_data['risk_factors']:
                        ws[f'B{row}'] = f"• {factor}"
                        row += 1
                    row += 2
        
        # Apply styles
        for cell in ws['A1:A' + str(ws.max_row)]:
            if cell[0].value and ':' not in str(cell[0].value):
                self._apply_style(cell[0], self.styles['metric_label'])

    def _add_interest_rate_risk_analysis(self, ws) -> None:
        """Add interest rate risk analysis"""
        row = ws.max_row + 3
        ws[f'A{row}'] = 'Interest Rate Risk Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2

        # Absorption Risk
        absorption_risk = {
            'Risk Assessment': f"Correlation coefficient of {self.market_analyzer._calculate_historical_rate_impact(self.market_analyzer.macro_data['interest_rates']['historical_trends']['5yr_fixed']):.2f} with rates",
            'Impact Analysis': [
                f"Base absorption at {self.market_analyzer.analysis_results['absorption_analysis']['current_rate']:.1f}% vs 65% target",
                f"Supply elasticity of {self.market_analyzer._calculate_supply_impact()*100:.1f}% per 10% supply increase",
                f"Rate sensitivity of {self.market_analyzer._calculate_rate_impact()*100:.1f}% per 100bps"
            ],
            'Risk Factors': [
                f"Manhattan (422 units) at 6 months absorption",
                f"Parkway 2 (396 units) at 12 months absorption",
                f"Standing inventory of {sum([74, 86, 64, 45])} units across active projects"
            ],
            'Mitigation Analysis': [
                'Structure release tranches based on absorption velocity differentials',
                'Implement dynamic pricing based on competitor absorption patterns',
                'Develop contingent incentive programs tied to market conditions'
            ]
        }
        self._add_risk_analysis_group(ws, row, 'Absorption Risk Analysis', absorption_risk)
        row += len(absorption_risk['Impact Analysis']) + len(absorption_risk['Risk Factors']) + 8

        # Price Risk Analysis
        price_risk = {
            'Risk Assessment': f"Price elasticity of {self.market_analyzer._calculate_price_premium(self.market_analyzer.analysis_results['pricing_analysis']['current_metrics']['avg_psf']):.2f}",
            'Impact Analysis': [
                f"Current market average PSF: ${self.market_analyzer.analysis_results['pricing_analysis']['current_metrics']['avg_psf']:.2f}",
                f"Premium decay pattern: {(1 - self.market_analyzer._calculate_rate_impact()) * 100:.1f}% over 12 months",
                f"Unit type spread: {self.market_analyzer.analysis_results['unit_type_analysis']['studios']['pricing_metrics']['avg_psf'] - self.market_analyzer.analysis_results['unit_type_analysis']['one_bed']['pricing_metrics']['avg_psf']:.2f} PSF"
            ],
            'Risk Factors': [
                'Manhattan establishing premium price points Oct 2024',
                'Parkway 2 validating market depth Apr 2024',
                'Completion timeline spread through 2025-2029'
            ],
            'Mitigation Analysis': [
                'Structure unit-specific beta-adjusted pricing model',
                'Implement dynamic pricing based on absorption velocity',
                'Develop strategic holds based on price sensitivity analysis'
            ]
        }
        self._add_risk_analysis_group(ws, row, 'Price Risk Analysis', price_risk)
        row += len(price_risk['Impact Analysis']) + len(price_risk['Risk Factors']) + 8

        # Supply Risk Analysis
        supply_risk = {
            'Risk Assessment': f"Pipeline analysis indicates {self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']['total_units']:,} units total",
            'Impact Analysis': [
                f"Active Projects: {self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']['active_units']:,} units",
                f"Standing Inventory: {self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']['standing_units']:,} units",
                f"Supply impact on absorption: {self.market_analyzer._calculate_supply_impact()*100:.1f}% per 10% increase"
            ],
            'Risk Factors': [
                'Active project absorption velocity',
                'Standing inventory levels',
                'Market maturity effect on pricing power'
            ],
            'Mitigation Analysis': [
                'Structure phased release strategy based on completion timeline',
                'Implement supply-adjusted pricing model by unit type',
                'Develop completion risk-based incentive programs'
            ]
        }
        self._add_risk_analysis_group(ws, row, 'Supply Risk Analysis', supply_risk)
        row += len(supply_risk['Impact Analysis']) + len(supply_risk['Risk Factors']) + 8

        # Market Risk Analysis
        market_risk = {
            'Risk Assessment': f"Composite risk score: {self.market_analyzer.analysis_results['market_score']:.1f}/10",
            'Impact Analysis': [
                f"Interest rate sensitivity: {self.market_analyzer._calculate_historical_rate_impact(self.market_analyzer.macro_data['interest_rates']['historical_trends']['5yr_fixed'])*100:.1f}% correlation",
                f"Employment elasticity: {self.market_analyzer._calculate_employment_impact(0.65):.2f}",
                f"Supply-adjusted beta: {self.market_analyzer._calculate_supply_impact() * 1.5:.2f}"
            ],
            'Risk Factors': [
                'Rate environment evolution through 2024-2025',
                'Employment trend impact on absorption velocity',
                'Market maturity effect on pricing power'
            ],
            'Mitigation Analysis': [
                'Optimize existing beta-adjusted pricing model with 2024 market data',
                'Enhance rate sensitivity calibration based on competitor performance',
                'Refine market condition-linked deposit programs'
            ]
        }
        self._add_risk_analysis_group(ws, row, 'Market Risk Analysis', market_risk)

    def _add_risk_analysis_group(self, ws, row: int, title: str, analysis: Dict) -> None:
        """Add sophisticated risk analysis group with detailed formatting"""
        ws[f'A{row}'] = title
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2

        # Add risk assessment
        ws[f'A{row}'] = 'Risk Assessment'
        ws[f'B{row}'] = analysis['Risk Assessment']
        self._apply_style(ws[f'A{row}'], self.styles['metric_label'])
        row += 2

        # Add impact analysis
        ws[f'A{row}'] = 'Impact Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        for impact in analysis['Impact Analysis']:
            ws[f'B{row}'] = f"• {impact}"
            row += 1
        row += 1

        # Add risk factors
        ws[f'A{row}'] = 'Risk Factors'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        for factor in analysis['Risk Factors']:
            ws[f'B{row}'] = f"• {factor}"
            row += 1
        row += 1

        # Add mitigation analysis
        ws[f'A{row}'] = 'Mitigation Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        for mitigation in analysis['Mitigation Analysis']:
            ws[f'B{row}'] = f"• {mitigation}"
            row += 1

    def _add_risk_mitigation_plan(self, ws) -> None:
        """Add sophisticated quantitative risk mitigation strategies based on our analysis"""
        row = ws.max_row + 5
        ws[f'A{row}'] = 'Quantitative Risk Mitigation Framework'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2

        # Updated Impact Analysis with correct metrics
        ws[f'A{row}'] = 'Impact Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        supply_data = self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']
        supply_impacts = [
            f"Active Projects: {supply_data['active_units']:,} units",
            f"Standing Inventory: {supply_data['standing_units']:,} units",
            f"Supply impact on absorption: {self.market_analyzer._calculate_supply_impact()*100:.1f}% per 10% increase"
        ]
        for impact in supply_impacts:
            ws[f'B{row}'] = f"• {impact}"
            row += 1
        row += 1

        # Market Competition Strategy
        competition_strategy = {
            'Risk Assessment': 'High - Launching into market with significant standing inventory',
            'Impact Analysis': f"Current standing inventory: Juno (74), Georgetown Two (86), Century City (64), Parkway 1 (45)",
            'Strategic Approach': [
                'Position against existing standing inventory in current selling projects',
                'Monitor absorption rates of current standing inventory through 2024-2025',
                'Track pricing evolution in projects under construction'
            ],
            'Monitoring Framework': [
                'Weekly absorption tracking of standing inventory in active projects',
                'Price point analysis of current selling inventory',
                'Construction progress monitoring of competitor projects'
            ]
        }
        self._add_sophisticated_metric_group(ws, row, 'Competition Strategy', competition_strategy)
        row += len(competition_strategy['Strategic Approach']) + len(competition_strategy['Monitoring Framework']) + 8

        # Supply Risk Management
        supply_strategy = {
            'Risk Assessment': 'High - Multiple projects in various construction stages',
            'Impact Analysis': f"Standing inventory of {sum([74, 86, 64, 45])} units across active projects",
            'Strategic Approach': [
                'Differentiate from current standing inventory pricing strategies',
                'Monitor completion timeline progression of under-construction projects',
                'Track pre-sale absorption rates of 2024 launches (Manhattan, Parkway 2)'
            ],
            'Monitoring Framework': [
                'Construction milestone tracking for all active projects',
                'Standing inventory absorption velocity analysis',
                'Pre-sale market response analysis for 2024 launches'
            ]
        }
        self._add_sophisticated_metric_group(ws, row, 'Supply Risk Management', supply_strategy)

    def _add_sophisticated_metric_group(self, ws, row: int, title: str, metrics: Dict) -> None:
        """Add sophisticated metric group with detailed formatting"""
        ws[f'A{row}'] = title
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2

        # Add risk assessment and impact analysis
        ws[f'A{row}'] = 'Risk Assessment'
        ws[f'B{row}'] = metrics['Risk Assessment']
        self._apply_style(ws[f'A{row}'], self.styles['metric_label'])
        row += 1

        ws[f'A{row}'] = 'Impact Analysis'  # Changed from 'Impact Quantification'
        ws[f'B{row}'] = metrics['Impact Analysis']  # Changed from 'Impact Quantification'
        self._apply_style(ws[f'A{row}'], self.styles['metric_label'])
        row += 2

        # Add strategies and considerations
        for category in ['Strategic Approach', 'Implementation', 'Monitoring Framework']:
            if category in metrics:
                ws[f'A{row}'] = category
                self._apply_style(ws[f'A{row}'], self.styles['subheader'])
                row += 1
                
                for item in metrics[category]:
                    ws[f'B{row}'] = f"• {item}"
                    row += 1
            row += 1

    def _determine_market_position(self) -> str:
        """Determine overall market positioning strategy"""
        market_score = self.market_analyzer.analysis_results['market_score']
        absorption_rate = self.market_analyzer.analysis_results['absorption_analysis']['current_rate']
        supply_data = self.market_analyzer.analysis_results['supply_analysis']
        
        # Calculate supply pressure using active units and total pipeline
        if isinstance(supply_data.get('supply_pressure'), dict):
            pressure_values = [v.get('absorption_pressure', 0) 
                             for v in supply_data['supply_pressure'].values()]
            supply_pressure = sum(pressure_values) / len(pressure_values) if pressure_values else 0.0
        else:
            # Calculate simple supply pressure from pipeline
            pipeline = supply_data['current_pipeline']
            total_units = pipeline['total_units']
            active_units = pipeline['active_units']
            supply_pressure = (active_units / total_units) if total_units > 0 else 0.0
        
        # Determine position based on market conditions
        if market_score >= 7.5 and absorption_rate >= 65 and supply_pressure < 1.2:
            return "Premium Market Leader - Opportunity for Aggressive Pricing"
        elif market_score >= 6.5 and absorption_rate >= 55 and supply_pressure < 1.5:
            return "Strong Market Position - Selective Premium Strategy"
        elif market_score >= 5.5 and absorption_rate >= 45:
            return "Competitive Market Position - Value-Based Pricing"
        else:
            return "Value-Focused Position - Competitive Pricing Required"

    def _determine_price_trend(self) -> str:
        """Determine price trend recommendation"""
        absorption_rate = self.market_analyzer.analysis_results['absorption_analysis']['current_rate']
        supply_data = self.market_analyzer.analysis_results['supply_analysis']
        
        # Calculate supply pressure from supply data
        supply_pressure = 0.0
        if isinstance(supply_data.get('supply_pressure'), dict):
            pressure_values = [v.get('absorption_pressure', 0) 
                             for v in supply_data['supply_pressure'].values()]
            supply_pressure = sum(pressure_values) / len(pressure_values) if pressure_values else 0.0
        else:
            # Calculate simple supply pressure from pipeline
            pipeline = supply_data['current_pipeline']
            total_units = pipeline['total_units']
            active_units = pipeline['active_units']
            supply_pressure = (active_units / total_units) if total_units > 0 else 0.0
        
        # Get interest rate trend
        interest_rates = self.market_analyzer.analysis_results['market_factors']['interest_rates']
        rate_trend = ("decreasing" 
                     if interest_rates['historical_trends']['5yr_fixed']['2024_avg'] < 
                        interest_rates['historical_trends']['5yr_fixed']['2023_avg'] 
                     else "increasing")
        
        # Get employment trend
        employment_trend = (self.market_analyzer.analysis_results['market_factors']['employment']
                           ['historical_trends']['2024_ytd']['employment_rate']['trend'])
        
        # Determine price trend based on multiple factors
        if (absorption_rate >= 70 and supply_pressure < 1.2 and 
            rate_trend == "decreasing" and employment_trend == "increasing"):
            return "Strong Support for Price Increases (3-5%)"
        elif (absorption_rate >= 55 and supply_pressure < 1.5 and 
              (rate_trend == "decreasing" or employment_trend == "increasing")):
            return "Moderate Price Growth Potential (2-3%)"
        elif absorption_rate >= 45:
            return "Stable Pricing with Selective Increases"
        else:
            return "Maintain Competitive Pricing"

    def _determine_competitive_position(self) -> Dict:
        """Determine competitive positioning strategy"""
        manhattan = self._get_competitor_data('The Manhattan')
        parkway2 = self._get_competitor_data('Parkway 2 - Intersect')
        
        manhattan_launch = datetime.strptime(manhattan['sales_start'], '%Y-%m-%d')
        parkway2_launch = datetime.strptime(parkway2['sales_start'], '%Y-%m-%d')
        
        return {
            'manhattan_strategy': self._determine_manhattan_strategy(manhattan),
            'parkway2_strategy': self._determine_parkway2_strategy(parkway2),
            'timing_advantage': self._analyze_timing_advantage(manhattan_launch, parkway2_launch),
            'positioning_strategy': self._determine_positioning_strategy()
        }

    def _determine_manhattan_strategy(self, manhattan: Dict) -> str:
        """Determine strategy relative to Manhattan"""
        manhattan_psf = np.mean([
            pricing['avg_psf'] 
            for pricing in manhattan['pricing'].values()
            if isinstance(pricing, dict) and 'avg_psf' in pricing
        ])
        if manhattan_psf > self.pricing_strategy['base_psf'] * 1.05:
            return "Position as Better Value Alternative"
        else:
            return "Match Quality with Competitive Pricing"

    def _determine_parkway2_strategy(self, parkway2: Dict) -> str:
        """Determine strategy relative to Parkway 2"""
        parkway2_psf = np.mean([
            pricing['avg_psf'] 
            for pricing in parkway2['pricing'].values()
            if isinstance(pricing, dict) and 'avg_psf' in pricing
        ])
        if parkway2_psf > self.pricing_strategy['base_psf'] * 1.03:
            return "Aggressive Competitive Positioning"
        else:
            return "Differentiate Through Product Features"

    def _analyze_timing_advantage(self, manhattan_launch: datetime, 
                                    parkway2_launch: datetime) -> str:
        """Analyze timing advantage/disadvantage"""
        our_launch = datetime.strptime('2024-04-01', '%Y-%m-%d')  # Example launch date
        
        manhattan_diff = (manhattan_launch - our_launch).days
        parkway2_diff = (parkway2_launch - our_launch).days
        
        if manhattan_diff > 90 and parkway2_diff > 45:
            return "Strong First-Mover Advantage"
        elif manhattan_diff > 60 or parkway2_diff > 30:
            return "Moderate Timing Advantage"
        else:
            return "Limited Timing Advantage"

    def _determine_positioning_strategy(self) -> str:
        """Determine overall positioning strategy"""
        market_score = self.market_analyzer.analysis_results['market_score']
        absorption_rate = self.market_analyzer.analysis_results['absorption_analysis']['current_rate']
        supply_data = self.market_analyzer.analysis_results['supply_analysis']
        
        if market_score >= 7.0 and absorption_rate >= 60:
            return "Premium Position with Market Leadership"
        elif market_score >= 6.0 and absorption_rate >= 50:
            return "Strong Position with Competitive Edge"
        else:
            return "Value Position with Quality Focus"

    def _add_completion_timeline_analysis(self, ws, row: int) -> int:
        """Add completion timeline analysis with quarterly breakdowns"""
        completions = {
            'The Manhattan': {'date': '2029-10-01', 'units': 422, 'status': 'Pre-Construction'},
            'Parkway 2': {'date': '2028-12-31', 'units': 396, 'status': 'Lot cleared'},
            'Juno': {'date': '2028-06-30', 'units': 341, 'status': 'Excavating'},
            'Sequoia': {'date': '2027-03-31', 'units': 386, 'status': 'Excavating'},
            'Georgetown': {'date': '2026-03-01', 'units': 355, 'status': 'Under Construction'},
            'Century Central': {'date': '2025-09-01', 'units': 409, 'status': 'Framing'},
            'Parkway 1': {'date': '2026-12-31', 'units': 363, 'status': 'Framing'}
        }
        
        ws[f'A{row}'] = 'Completion Timeline Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        # Create yearly summary
        yearly_summary = {}
        for project, details in completions.items():
            year = details['date'][:4]
            if year not in yearly_summary:
                yearly_summary[year] = {
                    'total_units': 0,
                    'projects': [],
                    'quarterly': {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
            }
            
            yearly_summary[year]['total_units'] += details['units']
            yearly_summary[year]['projects'].append(project)
            
            # Add to quarterly breakdown
            month = int(details['date'][5:7])
            quarter = f'Q{(month-1)//3 + 1}'
            yearly_summary[year]['quarterly'][quarter] += details['units']
        
        # Add yearly summary
        for year, data in sorted(yearly_summary.items()):
            ws[f'A{row}'] = f'{year} Completions'
            self._apply_style(ws[f'A{row}'], self.styles['subheader'])
            row += 1
            
            metrics = {
                'Total Units': f"{data['total_units']:,}",
                'Project Count': str(len(data['projects'])),
                'Projects': ', '.join(data['projects'])
            }
            
            self._add_metric_group(ws, row, 'Summary', metrics)
            row += len(metrics) + 1
            
            # Add quarterly breakdown
            quarterly_metrics = {
                f"{q}": f"{units:,} units"
                for q, units in data['quarterly'].items()
            }
            self._add_metric_group(ws, row, 'Quarterly Distribution', quarterly_metrics)
            row += len(quarterly_metrics) + 2
            
            # Add quarterly chart
            self._add_quarterly_chart(ws, f'D{row-6}', data['quarterly'], year)
            row += 15  # Space for chart
        
        return row

    def _add_quarterly_chart(self, ws, position: str, quarterly_data: Dict, year: str) -> None:
        """Add quarterly completion chart"""
        chart = BarChart()
        chart.title = f"{year} Quarterly Completions"
        chart.x_axis.title = "Quarter"
        chart.y_axis.title = "Units"
        
        # Add data
        row = ws.max_row + 2
        for quarter, units in quarterly_data.items():
            ws.cell(row=row, column=1, value=quarter)
            ws.cell(row=row, column=2, value=units)
            row += 1
        
        data = Reference(ws, min_col=2, min_row=row-4, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=row-4, max_row=row-1)
        
        chart.add_data(data)
        chart.set_categories(cats)
        
        ws.add_chart(chart, position)

    def _create_sensitivity_heatmaps(self, ws, sensitivity_data, row: int, label: str) -> int:
        """Create sensitivity heatmaps with improved color scaling"""
        color_scale = {
            'high': PatternFill(start_color='63BE7B', end_color='63BE7B', fill_type='solid'),  # Green
            'medium': PatternFill(start_color='FFEB84', end_color='FFEB84', fill_type='solid'), # Yellow
            'low': PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid')     # Red
        }

        current_row = row
        current_col = 2  # Start from column B to allow space for labels

        for key, data in sensitivity_data.items():
            if 'matrix' in data:
                # Add section header
                ws.cell(row=current_row, column=current_col, value=f"{label} Sensitivity: {key.replace('_vs_', ' vs ')}")
                self._apply_style(ws.cell(row=current_row, column=current_col), self.styles['subheader'])
                current_row += 2

                matrix = np.array(data['matrix'])
                
                # Add axis labels
                factors = key.split('_vs_')
                ws.cell(row=current_row-1, column=current_col, value=f"{factors[1]} →")
                ws.cell(row=current_row, column=1, value=f"{factors[0]} ")
                
                for i in range(len(matrix)):
                    for j in range(len(matrix[0])):
                        cell = ws.cell(row=current_row + i, column=current_col + j)
                        value = matrix[i,j]  # Already annual rate
                        cell.value = f"{value:.1f}%"
                        
                        # Apply color coding based on annual absorption thresholds
                        if value >= 60:  # ≥60% annual
                            cell.fill = color_scale['high']
                        elif value >= 42:  # 42-60% annual
                            cell.fill = color_scale['medium']
                        else:  # <42% annual
                            cell.fill = color_scale['low']

                current_row += len(matrix) + 3

        return current_row

    def _generate_sensitivity_findings(self, factor1: str, factor2: str, 
                                     matrix: List[List[float]], 
                                     values1: List[float], 
                                     values2: List[float]) -> Dict:
        """Generate findings from sensitivity analysis"""
        matrix_np = np.array(matrix)
        max_idx = np.unravel_index(matrix_np.argmax(), matrix_np.shape)
        min_idx = np.unravel_index(matrix_np.argmin(), matrix_np.shape)
        
        findings = {
            'Maximum Impact': (
                f"Peak absorption ({matrix_np[max_idx]:.1f}%) at "
                f"{factor1}: {values1[max_idx[0]]:.1f}%, "
                f"{factor2}: {values2[max_idx[1]]:.1f}%"
            ),
            'Minimum Impact': (
                f"Lowest absorption ({matrix_np[min_idx]:.1f}%) at "
                f"{factor1}: {values1[min_idx[0]]:.1f}%, "
                f"{factor2}: {values2[min_idx[1]]:.1f}%"
            ),
            'Sensitivity': self._determine_sensitivity_level(matrix_np, factor1, factor2),
            'Risk Implication': self._determine_risk_implication(matrix_np, factor1, factor2)
        }
        
        return findings

    def _determine_sensitivity_level(self, matrix: np.ndarray, 
                                       factor1: str, factor2: str) -> str:
        """Determine sensitivity level based on variance"""
        variance = np.var(matrix)
        
        if variance > 400:  # High variance in absorption rates
            return f"Highly sensitive to {factor1} and {factor2} changes"
        elif variance > 200:
            return f"Moderately sensitive to {factor1} and {factor2} changes"
        else:
            return f"Limited sensitivity to {factor1} and {factor2} changes"

    def _determine_risk_implication(self, matrix: np.ndarray, 
                                  factor1: str, factor2: str) -> str:
        """Determine risk implications from sensitivity analysis"""
        # Calculate percentage of scenarios below target absorption
        target_absorption = 65  # Target absorption rate
        below_target = np.sum(matrix < target_absorption) / matrix.size * 100
        
        if below_target > 66:
            return f"High risk - most scenarios below target absorption"
        elif below_target > 33:
            return f"Medium risk - mixed absorption outcomes"
        else:
            return f"Low risk - most scenarios achieve target absorption"

    def _calculate_unit_type_sensitivity(self, unit_type: str) -> Dict:
        """Calculate sensitivity matrices for specific unit type"""
        # Define sensitivity ranges
        factors = {
            'interest_rates': np.arange(-2.0, 2.1, 0.5),
            'price_change': np.arange(-10.0, 10.1, 2.5),
            'supply': np.arange(-20.0, 20.1, 5.0)
        }
        
        # Define base annual absorption rates by unit type
        base_annual_rates = {
            'studios': 66,      # 66% annual
            'one_bed': 70,      # 70% annual
            'two_bed': 65,      # 65% annual
            'three_bed': 55     # 55% annual
        }
        
        # Use the base rate for the unit type
        base_absorption = base_annual_rates[unit_type]
        
        # Calculate sensitivity matrices
        sensitivity_matrices = {}
        for factor1, values1 in factors.items():
            for factor2, values2 in factors.items():
                if factor1 < factor2:
                    matrix = np.zeros((len(values1), len(values2)))
                    for i, val1 in enumerate(values1):
                        for j, val2 in enumerate(values2):
                            matrix[i, j] = self._calculate_unit_impact(
                                unit_type,
                                base_absorption,
                                factor1, val1,
                                factor2, val2
                            )
                    sensitivity_matrices[f"{factor1}_vs_{factor2}"] = {
                        'matrix': matrix.tolist(),
                        'factor1_values': values1.tolist(),
                        'factor2_values': values2.tolist()
                    }
        
        return sensitivity_matrices

    def _calculate_unit_impact(self, unit_type: str, base_absorption: float,
                                 factor1: str, value1: float, 
                                 factor2: str, value2: float) -> float:
        """Calculate impact on unit type absorption with annual rates"""
        # Get unit type specific weights
        weights = {
            'studios': {
                'interest_rates': {
                    'up': -0.30,    # 30% reduction per 1% rate increase
                    'down': 0.12    # 12% increase per 1% rate decrease
                },
                'price_change': {
                    'up': -0.15,    # 15% reduction per 10% price increase
                    'down': 0.18    # 18% increase per 10% price decrease
                },
                'supply': -0.03     # 3% reduction per 10% supply increase
            },
            'one_bed': {
                'interest_rates': {
                    'up': -0.30,
                    'down': 0.12
                },
                'price_change': {
                    'up': -0.15,
                    'down': 0.18
                },
                'supply': -0.03
            },
            'two_bed': {
                'interest_rates': {
                    'up': -0.30,
                    'down': 0.12
                },
                'price_change': {
                    'up': -0.15,
                    'down': 0.18
                },
                'supply': -0.03
            },
            'three_bed': {
                'interest_rates': {
                    'up': -0.30,
                    'down': 0.12
                },
                'price_change': {
                    'up': -0.15,
                    'down': 0.18
                },
                'supply': -0.03
            }
        }
        
        unit_weights = weights[unit_type]
        
        # Calculate impacts with proper directional effects
        impact1 = 1.0
        if factor1 == 'interest_rates':
            if value1 > 0:
                # Base impact: 30% reduction per 1% increase
                impact1 = 1 - (value1 * 0.30)
                
                # Additional scaling for rates above 1%
                if value1 > 1.0:
                    excess = value1 - 1.0
                    # Progressive reduction: 40% for 1-1.5%, 60% for 1.5-2%
                    if excess <= 0.5:
                        impact1 *= (1 - (excess * 0.4))
                    else:
                        impact1 *= 0.8
                        impact1 *= (1 - ((excess - 0.5) * 0.8))
            else:
                # Positive impact for rate decreases
                impact1 = 1 + (abs(value1) * unit_weights['interest_rates']['down'])
        elif factor1 == 'price_change':
            if value1 > 0:
                # Negative impact for price increases
                impact1 = 1 - (value1/10 * abs(unit_weights['price_change']['up']))
            else:
                # Positive impact for price decreases
                impact1 = 1 + (abs(value1)/10 * unit_weights['price_change']['down'])
        else:  # supply
            impact1 = 1 + (value1/10 * unit_weights['supply'])
        
        impact2 = 1.0
        if factor2 == 'interest_rates':
            if value2 > 0:
                # Base impact: 30% reduction per 1% increase
                impact2 = 1 - (value2 * 0.30)
                
                # Additional scaling for rates above 1%
                if value2 > 1.0:
                    excess = value2 - 1.0
                    if excess <= 0.5:
                        impact2 *= (1 - (excess * 0.4))
                    else:
                        impact2 *= 0.8
                        impact2 *= (1 - ((excess - 0.5) * 0.8))
            else:
                # Positive impact for rate decreases
                impact2 = 1 + (abs(value2) * unit_weights['interest_rates']['down'])
        elif factor2 == 'price_change':
            if value2 > 0:
                # Negative impact for price increases
                impact2 = 1 - (value2/10 * abs(unit_weights['price_change']['up']))
            else:
                # Positive impact for price decreases
                impact2 = 1 + (abs(value2)/10 * unit_weights['price_change']['down'])
        else:  # supply
            impact2 = 1 + (value2/10 * unit_weights['supply'])
        
        # Calculate final annual absorption
        final_absorption = base_absorption * impact1 * impact2
        
        # Return annual rate with appropriate bounds
        return max(15.0, min(85.0, final_absorption))

    def _calculate_unit_price_points(self, unit_type: str, base_psf: float) -> Dict:
        """Calculate price points for unit type"""
        # Use correct market PSF from analysis
        market_psf = self.market_analyzer.analysis_results['pricing_analysis']['current_metrics']['avg_psf']  # This is 1187
        
        # Define target size ranges for each unit type
        size_ranges = {
            'studios': [300, 400],
            'one_bed': [450, 550],
            'two_bed': [700, 850],
            'three_bed': [900, 1100]
        }
        
        # Premium/discount structure (keeping same as before)
        premiums = {
            'studios': 5.0,      # +5.0% premium
            'one_bed': 2.0,      # +2.0% premium
            'two_bed': -3.0,     # -3.0% discount
            'three_bed': -6.0    # -6.0% discount
        }
        
        # Calculate target PSF with premium
        premium = premiums[unit_type]
        target_psf = market_psf * (1 + premium/100)
        
        # Calculate price range based on size range
        min_size, max_size = size_ranges[unit_type]
        min_price = round(min_size * target_psf / 5000) * 5000  # Round to nearest $5k
        max_price = round(max_size * target_psf / 5000) * 5000
        
        # Generate appropriate strategy message
        strategy_messages = {
            'studios': f"Premium pricing at {premium:+.1f}% reflecting highest efficiency and strong absorption at 11.8%",
            'one_bed': f"Strong market positioning at {premium:+.1f}% premium with 3.5% absorption",
            'two_bed': f"Competitive pricing at {premium:+.1f}% to market targeting end-user demand",
            'three_bed': f"Value-oriented pricing at {premium:+.1f}% for larger units with 11.8% absorption"
        }
        
        return {
            'base_psf': market_psf,
            'premium': premium,
            'target_psf': target_psf,
            'min_price': min_price,
            'max_price': max_price,
            'strategy': strategy_messages[unit_type]
        }

    def _calculate_optimal_premium(self, unit_type: str, absorption_rate: Dict, 
                                 demand_index: float, competitor_psf_manhattan: float,
                                 competitor_psf_parkway: float, 
                                 sensitivity_data: Dict) -> float:
        """Calculate optimal premium based on comprehensive analysis"""
        # Extract monthly absorption rate from dictionary
        monthly_absorption = absorption_rate['monthly']
        
        # 1. Base premium from unit type efficiency (smaller units command higher PSF)
        efficiency_premiums = {
            'studios': 6.0,     # Highest efficiency premium due to smallest size
            'one_bed': 3.0,     # Strong efficiency but less than studios
            'two_bed': -2.0,    # Slight discount due to size
            'three_bed': -4.0   # Larger discount for largest units
        }
        base_premium = efficiency_premiums[unit_type]
        
        # 2. Demand-based adjustment
        demand_adjustment = (demand_index - 1.0) * 3.0  # +/- 3% per 1.0 demand index difference
        
        # 3. Absorption-based adjustment
        absorption_adjustment = 0.0
        if monthly_absorption > 65:
            absorption_adjustment = 2.0
        elif monthly_absorption < 45:
            absorption_adjustment = -2.0
        
        # 4. Competitive positioning adjustment
        comp_avg_psf = np.mean([p for p in [competitor_psf_manhattan, competitor_psf_parkway] if p > 0])
        if comp_avg_psf > 0:
            competitive_gap = 2.0  # Target slight discount to competitors
        
        # 5. Risk adjustment from sensitivity analysis
        risk_adjustment = self._calculate_risk_based_adjustment(unit_type, sensitivity_data)
        
        # Calculate final premium
        final_premium = (
            base_premium +          # Base efficiency premium
            demand_adjustment +     # Demand-driven adjustment
            absorption_adjustment + # Absorption-based adjustment
            risk_adjustment        # Risk-based adjustment
        )
        
        # Cap maximum premium/discount based on unit type
        premium_caps = {
            'studios': (5.0, 10.0),    # Min 5%, Max 10%
            'one_bed': (2.0, 6.0),     # Min 2%, Max 6%
            'two_bed': (-3.0, 0.0),    # Min -3%, Max 0%
            'three_bed': (-6.0, -3.0)  # Min -6%, Max -3%
        }
        
        min_premium, max_premium = premium_caps[unit_type]
        final_premium = max(min_premium, min(max_premium, final_premium))

    def _calculate_risk_based_adjustment(self, unit_type: str, sensitivity_data: Dict) -> float:
        """Calculate risk-based price adjustment"""
        # Extract price sensitivity from analysis
        price_sensitivity = next(
            (data for key, data in sensitivity_data.items() 
             if 'price_change' in key),
            None
        )
        
        if not price_sensitivity:
            return 0.0
        
        # Calculate risk metrics
        matrix = np.array(price_sensitivity['matrix'])
        risk_score = np.std(matrix) / np.mean(matrix)  # Coefficient of variation
        
        # More conservative pricing for higher risk
        if risk_score > 0.3:
            return -1.0  # High risk - reduce premium
        elif risk_score > 0.2:
            return -0.5  # Moderate risk - slight reduction
        return 0.0      # Low risk - no adjustment

    def _get_pricing_strategy(self, unit_type: str, premium: float, absorption_rate: Dict) -> str:
        """Get detailed pricing strategy based on premium and performance"""
        # Extract monthly absorption rate
        monthly_absorption = absorption_rate['monthly']
        
        if unit_type == 'studios':
            return (f"Premium pricing at {premium:+.1f}% reflecting highest efficiency "
                    f"and strong absorption at {monthly_absorption:.1f}%")
        elif unit_type == 'one_bed':
            return (f"Strong market positioning at {premium:+.1f}% premium "
                    f"with {monthly_absorption:.1f}% absorption")
        elif unit_type == 'two_bed':
            return (f"Competitive pricing at {premium:+.1f}% to market "
                    f"targeting end-user demand")
        else:  # three_bed
            return (f"Value-oriented pricing at {premium:+.1f}% for larger units "
                    f"with {monthly_absorption:.1f}% absorption")

    def generate_pdf_report(self, output_file: str = 'Surrey_Market_Analysis.pdf') -> None:
        """Generate comprehensive PDF report"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
                PageBreak, KeepTogether
            )
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
            from reportlab.graphics.charts.barcharts import VerticalBarChart
            
            # Create document with landscape orientation
            doc = SimpleDocTemplate(
                output_file,
                pagesize=landscape(letter),
                rightMargin=36,
                leftMargin=36,
                topMargin=36,
                bottomMargin=36
            )
            
            # Create styles
            styles = self._create_pdf_styles()
            
            # Build document elements
            elements = []
            
            # 1. Executive Summary
            elements.extend(self._create_executive_summary_pdf(styles))
            elements.append(PageBreak())
            
            # 2. Market Analysis (Presale Summary)
            elements.extend(self._create_market_analysis_pdf(styles))
            elements.append(PageBreak())
            
            # 3. Macro Landscape
            elements.extend(self._create_macro_landscape_pdf(styles))
            elements.append(PageBreak())
            
            # 4. Market Sensitivity Analysis
            elements.extend(self._create_market_sensitivity_pdf(styles))
            elements.append(PageBreak())
            
            # 5. Unit Sensitivity Analysis
            elements.extend(self._create_unit_sensitivity_pdf(styles))
            elements.append(PageBreak())
            
            # 6-10. Revenue Analysis (3M, 6M, 12M, 24M, 36M)
            periods = ['3_month', '6_month', '12_month', '24_month', '36_month']
            for period in periods:
                elements.extend(self._create_revenue_analysis_pdf(styles, period))
                elements.append(PageBreak())
            
            # Build PDF
            doc.build(elements)
            print(f"PDF report generated successfully: {output_file}")
            
        except Exception as e:
            print(f"Error generating PDF report: {str(e)}")
            traceback.print_exc()

    def _add_pdf_executive_summary(self, elements, RLParagraph, styles, heading_style):
        """Add executive summary section to PDF"""
        elements.append(RLParagraph("Executive Summary", heading_style))
        elements.append(Spacer(1, 12))
        
        # Key metrics
        data = [
            ['Metric', 'Value'],
            ['Market Score', f"{self.market_analyzer.analysis_results['market_score']:.1f}/10"],
            ['Current Absorption', f"{self.market_analyzer.analysis_results['absorption_analysis']['current_rate']:.1f}%"],
            ['Average PSF', f"${self.market_analyzer.analysis_results['pricing_analysis']['current_metrics']['avg_psf']:.2f}"],
            ['Supply Pipeline', f"{self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']['total_pipeline']:,} units"]
        ]
        
        table = Table(data, colWidths=[200, 200])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))    

    def _add_pdf_market_analysis(self, elements, RLParagraph, styles, heading_style):
        """Add market analysis section to PDF"""
        elements.append(RLParagraph("Market Analysis", heading_style))
        elements.append(Spacer(1, 12))
        
        # Add market metrics and analysis
        market_text = [
            f"The market demonstrates a score of {self.market_analyzer.analysis_results['market_score']:.1f}/10, ",
            f"with current absorption at {self.market_analyzer.analysis_results['absorption_analysis']['current_rate']:.1f}%. ",
            "Key findings include:"
        ]
        
        elements.append(RLParagraph(" ".join(market_text), styles['Normal']))
        elements.append(Spacer(1, 12))

    def _add_pdf_sensitivity_analysis(self, elements, RLParagraph, styles, heading_style, Image):
        """Add sensitivity analysis section with heatmaps to PDF"""
        elements.append(RLParagraph("Sensitivity Analysis", heading_style))
        elements.append(Spacer(1, 12))
        
        # Overall Market Sensitivity
        elements.append(RLParagraph("Market Sensitivity", styles['Heading3']))
        elements.append(Spacer(1, 6))
        
        # Create and add sensitivity heatmaps
        sensitivity_data = self.market_analyzer.analysis_results['sensitivity_analysis']
        for factor_combo, data in sensitivity_data.items():
            factors = factor_combo.split('_vs_')
            
            # Create heatmap
            plt.figure(figsize=(8, 6))
            sns.heatmap(
                np.array(data['matrix']),
                annot=True,
                fmt='.1f',
                cmap='RdYlGn',
                center=65,
                xticklabels=[f"{x:.1f}%" for x in data['factor2_values']],
                yticklabels=[f"{x:.1f}%" for x in data['factor1_values']]
            )
            
            plt.title(f'Market Sensitivity: {factors[0].title()} vs {factors[1].title()}')
            plt.xlabel(f'{factors[1].title()} Change (%)')
            plt.ylabel(f'{factors[0].title()} Change (%)')
            
            # Save heatmap
            temp_file = os.path.join(self.temp_dir, f'sensitivity_{factor_combo}.png')
            plt.savefig(temp_file, bbox_inches='tight', dpi=300)
            plt.close()
            
            # Add to PDF
            elements.append(Image(temp_file, width=400, height=300))
            elements.append(Spacer(1, 12))
            
            # Add findings
            findings = self._generate_sensitivity_findings(
                factors[0], factors[1], 
                data['matrix'],
                data['factor1_values'],
                data['factor2_values']
            )
            
            findings_data = [[k, v] for k, v in findings.items()]
            table = Table(findings_data, colWidths=[150, 300])
            table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('PADDING', (0, 0), (-1, -1), 6)
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
        
        # Add unit type sensitivity
        elements.append(RLParagraph("Unit Type Sensitivity", styles['Heading3']))
        elements.append(Spacer(1, 6))
        
        for unit_type in ['studios', 'one_bed', 'two_bed', 'three_bed']:
            elements.append(RLParagraph(
                f"{unit_type.replace('_', ' ').title()} Sensitivity", 
                styles['Heading4']
            ))
            elements.append(Spacer(1, 6))
            
            sensitivity = self._calculate_unit_type_sensitivity(unit_type)
            
            # Create and add unit type sensitivity heatmaps
            for factor_combo, data in sensitivity.items():
                factors = factor_combo.split('_vs_')
                
                plt.figure(figsize=(8, 6))
                sns.heatmap(
                    np.array(data['matrix']),
                    annot=True,
                    fmt='.1f',
                    cmap='RdYlGn',
                    center=65,
                    xticklabels=[f"{x:.1f}%" for x in data['factor2_values']],
                    yticklabels=[f"{x:.1f}%" for x in data['factor1_values']]
                )
                
                plt.title(f'{unit_type.title()} Sensitivity: {factors[0].title()} vs {factors[1].title()}')
                plt.xlabel(f'{factors[1].title()} Change (%)')
                plt.ylabel(f'{factors[0].title()} Change (%)')
                
                temp_file = os.path.join(self.temp_dir, f'sensitivity_{unit_type}_{factor_combo}.png')
                plt.savefig(temp_file, bbox_inches='tight', dpi=300)
                plt.close()
                
                elements.append(Image(temp_file, width=400, height=300))
                elements.append(Spacer(1, 12))
            
            elements.append(Spacer(1, 20))

    def _add_pdf_competitive_analysis(self, elements, RLParagraph, styles, heading_style):
        """Add competitive analysis section to PDF"""
        elements.append(RLParagraph("Competitive Analysis", heading_style))
        elements.append(Spacer(1, 12))
        
        # Add competitor analysis
        competitors = {
            'Manhattan': self._get_competitor_data('The Manhattan'),
            'Parkway 2': self._get_competitor_data('Parkway 2 - Intersect')
        }
        
        for name, data in competitors.items():
            comp_data = [
                ['Metric', 'Value'],
                ['Total Units', str(data['total_units'])],
                ['Absorption', f"{data['current_absorption']:.1f}%"],
                ['Launch Date', data['sales_start']]]
        
            table = Table(comp_data, colWidths=[200, 200])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(RLParagraph(f"{name} Analysis", styles['Heading3']))
            elements.append(Spacer(1, 6))
            elements.append(table)
            elements.append(Spacer(1, 12))

    def _add_pdf_recommendations(self, elements, RLParagraph, styles, heading_style):
        """Add recommendations section to PDF"""
        elements.append(RLParagraph("Recommendations", heading_style))
        elements.append(Spacer(1, 12))
        
        # Add pricing recommendations
        pricing_data = [
            ['Unit Type', 'Target PSF', 'Premium/Discount'],
        ]
        
        for unit_type in ['studios', 'one_bed', 'two_bed', 'three_bed']:
            price_points = self.market_analyzer.analysis_results['pricing_analysis']['unit_type_analysis'][unit_type]
            pricing_data.append([
                unit_type.replace('_', ' ').title(),
                f"${price_points['target_psf']:.2f}",
                f"{price_points['premium']:+.1f}%"
            ])
        
        table = Table(pricing_data, colWidths=[150, 150, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))

    def _add_revenue_analysis(self, ws, target_period: str, target_absorption: float) -> None:
        """Add revenue analysis sheet with absorption targets"""
        # Add title
        ws['A1'] = f'Revenue Analysis - Target {target_absorption}% Absorption'
        self._apply_style(ws['A1'], self.styles['header'])
        
        # Add monthly absorption table on top right
        self._add_monthly_absorption_table(ws, target_period, target_absorption)
        
        # Add gross revenue table
        row = 3
        ws[f'A{row}'] = 'Gross Revenue Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 2
        
        # Add headers
        headers = ['Unit Type', 'Units', '% Total', '$ Volume', '% Total $ Volume', 
                  'Total SF', 'Incentive', 'Avg PSF', 'Avg Size']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self._apply_style(ws.cell(row=row, column=col), self.styles['subheader'])
        
        # Get revenue metrics for the specific target period
        revenue_metrics = self.market_analyzer.analysis_results['revenue_metrics'][target_period]['unit_metrics']
        total_volume = self.market_analyzer.analysis_results['revenue_metrics'][target_period]['total_volume']
        total_sf = self.market_analyzer.analysis_results['revenue_metrics'][target_period]['total_sf']
        
        # Add data for each unit type
        row += 1
        for unit_type, metrics in revenue_metrics.items():
            col = 1
            ws.cell(row=row, column=col, value=unit_type.replace('_', ' ').title()); col += 1
            ws.cell(row=row, column=col, value=metrics['units']); col += 1
            ws.cell(row=row, column=col, value=f"{metrics['pct_total_units']:.1f}%"); col += 1
            ws.cell(row=row, column=col, value=metrics['gross_revenue']); col += 1
            ws.cell(row=row, column=col, value=f"{metrics['pct_total_volume']:.1f}%"); col += 1
            ws.cell(row=row, column=col, value=metrics['total_sf']); col += 1
            ws.cell(row=row, column=col, value=f"{metrics['incentive_rate']*100:.1f}%"); col += 1
            ws.cell(row=row, column=col, value=f"${metrics['target_psf']:.2f}"); col += 1
            ws.cell(row=row, column=col, value=metrics['avg_size'])
            row += 1
        
        # Add totals row
        row += 1
        ws.cell(row=row, column=1, value='Total')
        ws.cell(row=row, column=2, value=sum(m['units'] for m in revenue_metrics.values()))
        ws.cell(row=row, column=3, value='100.0%')
        ws.cell(row=row, column=4, value=total_volume)
        ws.cell(row=row, column=5, value='100.0%')
        ws.cell(row=row, column=6, value=total_sf)
        
        # Add net revenue table
        row += 3
        ws[f'A{row}'] = 'Net Revenue Analysis (After Incentives)'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 2
        
        # Add headers again
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self._apply_style(ws.cell(row=row, column=col), self.styles['subheader'])
        
        # Add net revenue data
        row += 1
        net_total_volume = 0
        for unit_type, metrics in revenue_metrics.items():
            net_revenue = metrics['gross_revenue'] * (1 - metrics['incentive_rate'])
            net_total_volume += net_revenue
            
            col = 1
            ws.cell(row=row, column=col, value=unit_type.replace('_', ' ').title()); col += 1
            ws.cell(row=row, column=col, value=metrics['units']); col += 1
            ws.cell(row=row, column=col, value=f"{metrics['pct_total_units']:.1f}%"); col += 1
            ws.cell(row=row, column=col, value=net_revenue); col += 1
            ws.cell(row=row, column=col, value=f"{(net_revenue/net_total_volume*100):.1f}%"); col += 1
            ws.cell(row=row, column=col, value=metrics['total_sf']); col += 1
            ws.cell(row=row, column=col, value=f"{metrics['incentive_rate']*100:.1f}%"); col += 1
            ws.cell(row=row, column=col, value=f"${metrics['net_psf']:.2f}"); col += 1
            ws.cell(row=row, column=col, value=metrics['avg_size'])
            row += 1
        
        # Add net totals row
        row += 1
        ws.cell(row=row, column=1, value='Total')
        ws.cell(row=row, column=2, value=sum(m['units'] for m in revenue_metrics.values()))
        ws.cell(row=row, column=3, value='100.0%')
        ws.cell(row=row, column=4, value=net_total_volume)
        ws.cell(row=row, column=5, value='100.0%')
        ws.cell(row=row, column=6, value=total_sf)

    def _add_monthly_absorption_table(self, ws, target_period: str, target_absorption: float) -> None:
        """Add monthly absorption and pricing strategy table"""
        try:
            start_col = 14
            
            # Add table header
            ws[f'{chr(64+start_col)}1'] = 'Monthly Absorption & Pricing Strategy'
            self._apply_style(ws[f'{chr(64+start_col)}1'], self.styles['subheader'])
            
            # Add headers
            headers = ['Month', 'Target %', 'Cumulative %']
            for i, header in enumerate(headers):
                ws.cell(row=2, column=start_col+i-2, value=header)
                self._apply_style(ws.cell(row=2, column=start_col+i-2), self.styles['subheader'])
            
            # Add unit type headers
            unit_types = ['Studios', 'One Bed', 'Two Bed', 'Three Bed']
            for i, unit_type in enumerate(unit_types):
                col = start_col + (i * 3)
                ws.cell(row=2, column=col, value=unit_type)
                ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+2)
                self._apply_style(ws.cell(row=2, column=col), self.styles['subheader'])
                
                # Sub-headers
                ws.cell(row=3, column=col, value='Gross PSF')
                ws.cell(row=3, column=col+1, value='Incentive')
                ws.cell(row=3, column=col+2, value='Net PSF')
                for subcol in range(3):
                    self._apply_style(ws.cell(row=3, column=col+subcol), self.styles['subheader'])
            
            # Get period-specific months
            period_months = {
                '3_month': 3,
                '6_month': 6,
                '12_month': 12,
                '24_month': 24,
                '36_month': 36
            }[target_period]
            
            # Get monthly pricing schedule
            monthly_pricing = self.market_analyzer._calculate_monthly_pricing_schedule()
            
            # Add data only for relevant months
            row = 4
            cumulative = 0
            for month in range(period_months):
                actual_month = ((4 + month - 1) % 12) + 1  # Start from April
                
                # Get monthly target
                month_target = monthly_pricing[month]['target_absorption']
                cumulative += month_target
                
                # Add absorption data
                ws.cell(row=row, column=start_col-2, value=datetime(2024, actual_month, 1).strftime('%B'))
                ws.cell(row=row, column=start_col-1, value=f"{month_target:.2f}%")
                ws.cell(row=row, column=start_col, value=f"{cumulative:.2f}%")
                
                # Add pricing data for each unit type
                for i, unit_type in enumerate(['studios', 'one_bed', 'two_bed', 'three_bed']):
                    pricing = monthly_pricing[month]['unit_types'][unit_type.replace('-', '_')]
                    col = start_col + (i * 3)
                    
                    ws.cell(row=row, column=col, value=f"${pricing['gross_psf']:.2f}")
                    ws.cell(row=row, column=col+1, value=f"{pricing['incentive']*100:.1f}%")
                    ws.cell(row=row, column=col+2, value=f"${pricing['net_psf']:.2f}")
                
                row += 1
                
        except Exception as e:
            print(f"Error adding absorption table: {str(e)}")

    def _calculate_weighted_psf(self, strategies: Dict, month: int) -> Dict:
        """Calculate weighted average PSF across unit types"""
        unit_weights = {
            'studios': 34/376,    # 9%
            'one_bed': 204/376,   # 54.3%
            'two_bed': 120/376,   # 31.9%
            'three_bed': 18/376   # 4.8%
        }
        
        weighted_gross = 0
        weighted_net = 0
        weighted_incentive = 0
        
        for unit_type, weight in unit_weights.items():
            strategy = strategies[unit_type][month]
            weighted_gross += strategy['gross_psf'] * weight
            weighted_net += strategy['net_psf'] * weight
            weighted_incentive += strategy['incentive_rate'] * weight
        
        return {
            'gross_psf': weighted_gross,
            'net_psf': weighted_net,
            'incentive_rate': weighted_incentive
        }

    def _get_competitor_data(self, project_name: str) -> Dict:
        """Get competitor data from market analysis"""
        try:
            # Get project data from active projects
            project = next((p for p in self.market_analyzer.project_data['active_projects']['projects'] 
                           if p['name'] == project_name), None)
            
            if project:
                # Calculate absorption
                total_units = project.get('total_units', 0)
                units_sold = project.get('units_sold', 0)
                current_absorption = (units_sold / total_units * 100) if total_units > 0 else 0
                
                # Get pricing data
                pricing_data = {}
                for unit_type in ['studios', 'one_bed', 'two_bed', 'three_bed']:
                    if unit_type in project.get('pricing', {}):
                        pricing_data[unit_type] = project['pricing'][unit_type]
                
                return {
                    'name': project_name,
                    'total_units': total_units,
                    'units_sold': units_sold,
                    'current_absorption': current_absorption,
                    'sales_start': project.get('sales_start', 'N/A'),
                    'completion': project.get('completion', 'N/A'),
                    'pricing': pricing_data
                }
            
            # Return empty data if project not found
            return {
                'name': project_name,
                'total_units': 0,
                'units_sold': 0,
                'current_absorption': 0.0,
                'sales_start': 'N/A',
                'completion': 'N/A',
                'pricing': {}
            }
            
        except Exception as e:
            print(f"Error getting competitor data for {project_name}: {str(e)}")
            return {
                'name': project_name,
                'total_units': 0,
                'units_sold': 0,
                'current_absorption': 0.0,
                'sales_start': 'N/A',
                'completion': 'N/A',
                'pricing': {}
            }

    def _create_recommendations(self) -> None:
        """Create recommendations sheet"""
        ws = self.wb.create_sheet("Recommendations")
        row = 1
        
        # Get recommendations from market analyzer
        recommendations = self.market_analyzer._generate_market_entry_recommendations()
        
        # Market Entry Strategy
        ws[f'A{row}'] = 'Market Entry Strategy'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        # Market Timing Analysis
        timing = recommendations['market_timing_analysis']
        ws[f'A{row}'] = 'Market Timing Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        ws[f'A{row}'] = 'Strategy'
        ws[f'B{row}'] = timing['strategy']
        row += 2
        
        ws[f'A{row}'] = 'Analysis'
        row += 1
        for point in timing['analysis']:
            ws[f'B{row}'] = f"• {point}"
            row += 1
        row += 1
        
        ws[f'A{row}'] = 'Implementation'
        row += 1
        for point in timing['implementation']:
            ws[f'B{row}'] = f"• {point}"
            row += 1
        row += 3  # Extra space between sections
        
        # Pricing Optimization
        pricing = recommendations['pricing_optimization']
        ws[f'A{row}'] = 'Pricing Optimization'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        ws[f'A{row}'] = 'Strategy'
        ws[f'B{row}'] = pricing['strategy']
        row += 2
        
        ws[f'A{row}'] = 'Analysis'
        row += 1
        for point in pricing['analysis']:
            ws[f'B{row}'] = f"• {point}"
            row += 1
        row += 1
        
        ws[f'A{row}'] = 'Implementation'
        row += 1
        for point in pricing['implementation']:
            ws[f'B{row}'] = f"• {point}"
            row += 1
        row += 3  # Extra space between sections
        
        # Market Positioning
        positioning = recommendations['market_positioning']
        ws[f'A{row}'] = 'Market Positioning'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        ws[f'A{row}'] = 'Strategy'
        ws[f'B{row}'] = positioning['strategy']
        row += 2
        
        ws[f'A{row}'] = 'Analysis'
        row += 1
        for point in positioning['analysis']:
            ws[f'B{row}'] = f"• {point}"
            row += 1
        row += 1
        
        ws[f'A{row}'] = 'Implementation'
        row += 1
        for point in positioning['implementation']:
            ws[f'B{row}'] = f"• {point}"
            row += 1
        row += 3  # Extra space between sections
        
        # Risk Mitigation
        risk = recommendations['risk_mitigation']
        ws[f'A{row}'] = 'Risk Mitigation'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        ws[f'A{row}'] = 'Strategy'
        ws[f'B{row}'] = risk['strategy']
        row += 2
        
        ws[f'A{row}'] = 'Analysis'
        row += 1
        for point in risk['analysis']:
            ws[f'B{row}'] = f"• {point}"
            row += 1
        row += 1
        
        ws[f'A{row}'] = 'Implementation'
        row += 1
        for point in risk['implementation']:
            ws[f'B{row}'] = f"• {point}"
            row += 1
        row += 5  # Large space before product strategy
        
        # Product Strategy (moved to bottom with clear separation)
        ws[f'A{row}'] = 'Product Strategy'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        # Unit Mix Performance
        ws[f'A{row}'] = 'Unit Mix Performance'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        # Create properly structured unit mix data
        unit_mix_data = {
            'Strategy': 'Optimize release based on absorption data',
            'Analysis': [
                f"One beds showing {self.market_analyzer.analysis_results['unit_type_analysis']['one_bed']['inventory_metrics']['absorption_rate']['monthly']:.1f}% monthly absorption",
                f"Two beds at {self.market_analyzer.analysis_results['unit_type_analysis']['two_bed']['inventory_metrics']['absorption_rate']['monthly']:.1f}% monthly absorption",
                f"Studios at {self.market_analyzer.analysis_results['unit_type_analysis']['studios']['inventory_metrics']['absorption_rate']['monthly']:.1f}% monthly absorption"
            ],
            'Implementation': [
                'Prioritize one bed release (54.3%) given proven absorption',
                'Balance two bed inventory (31.9%) release with Manhattan/Parkway 2 performance',
                'Strategic release of studios (9%) based on competitor data'
            ]
        }
        
        # Add strategy
        ws[f'A{row}'] = 'Strategy'
        ws[f'B{row}'] = unit_mix_data['Strategy']
        row += 2
        
        # Add analysis points
        ws[f'A{row}'] = 'Analysis'
        row += 1
        for point in unit_mix_data['Analysis']:
            ws[f'B{row}'] = f"• {point}"
            row += 1
        row += 1
        
        # Add implementation points
        ws[f'A{row}'] = 'Implementation'
        row += 1
        for point in unit_mix_data['Implementation']:
            ws[f'B{row}'] = f"• {point}"
            row += 1
        
        # Apply styles
        for cell in ws[f'A{row-len(unit_mix_data["Implementation"])-len(unit_mix_data["Analysis"])-4}:A{row}']:
            if cell[0].value in ['Strategy', 'Analysis', 'Implementation']:
                self._apply_style(cell[0], self.styles['metric_label'])
        
        return row

    def _create_executive_summary_pdf(self, styles) -> List:
        """Create executive summary section for PDF"""
        elements = []
        
        # Title
        elements.append(RLParagraph("Executive Summary", styles['Heading1']))
        elements.append(Spacer(1, 20))
        
        # 1. Market Overview
        elements.append(RLParagraph("Market Overview", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        market_metrics = [
            ['Metric', 'Value', 'Impact'],
            ['Active Project Units', 
             f"{self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']['active_units']:,}",
             "Current market supply"],
            ['Standing Inventory',
             f"{self.market_analyzer.analysis_results['supply_analysis']['current_pipeline']['standing_units']:,}",
             f"{self.market_analyzer._calculate_months_of_supply():.1f} months supply"],
            ['Current Absorption',
             f"{self.market_analyzer.analysis_results['absorption_analysis']['current_rate']:.1f}%",
             "Monthly absorption rate"],
            ['Average Price PSF',
             f"${self.market_analyzer.analysis_results['pricing_analysis']['current_metrics']['avg_psf']:,.0f}",
             "Market average"]
        ]
        
        # Create table with proper column widths and word wrapping
        table = Table(market_metrics, colWidths=[120, 100, 200])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),  # Smaller font size
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 35))
        
        # 2. Risk Analysis
        elements.append(RLParagraph("Risk Analysis", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        risk_analysis = self.simulation_results['risk_analysis']
        risk_metrics = [
            ['Risk Factor', 'Assessment', 'Impact', 'Mitigation'],
            ['Price Risk', 
             'Moderate',
             '±5% price change impacts\nabsorption by 10-15%',
             'Beta-adjusted pricing\nstrategy'],
            ['Interest Rate Risk',
             'High',
             'A 1% rate decrease increases\nabsorption by 14%',
             'Dynamic incentive model\naligned with rate changes'],
            ['Supply Risk',
             'Moderate',
             '10% supply increase reduces\nabsorption by 5%',
             'Phased release\nstrategy'],
            ['Market Position',
             'Strong',
             'Premium positioning validated\nby Manhattan/Parkway 2',
             'Maintain competitive\npricing with premiums']
        ]
        
        # Increased column widths and added word wrapping
        table = Table(risk_metrics, colWidths=[80, 70, 160, 140])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')  # Vertical alignment
        ]))
        elements.append(table)
        elements.append(Spacer(1, 50))
        
        # 3. Key Recommendations
        elements.append(RLParagraph("Key Recommendations", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        recommendations = [
            ['Category', 'Strategy', 'Implementation'],
            ['Pricing', 
             'Premium positioning with\nstrategic incentives',
             '• Base PSF aligned with Manhattan/Parkway 2\n• Dynamic incentives based on absorption\n• Unit-specific premium structure'],
            ['Absorption',
             'Phased release\nstrategy',
             '• 50% absorption in first 3 months\n• 65% by month 12\n• 82.5% by month 24\n• 100% by month 36'],
            ['Market Position',
             'Value-add\nopportunities',
             '• Premium finishes package\n• Flexible deposit structure\n• Strategic unit mix optimization'],
            ['Risk Management',
             'Dynamic response\nframework',
             '• Monthly absorption monitoring\n• Quarterly pricing reviews\n• Supply-based release strategy']
        ]
        
        # Increased column widths for recommendations
        table = Table(recommendations, colWidths=[90, 130, 230])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        elements.append(table)
        
        return elements

    def _create_market_analysis_pdf(self, styles) -> List:
        """Create market analysis section for PDF"""
        elements = []
        
        # Title
        elements.append(RLParagraph("Market Analysis", styles['Heading1']))
        elements.append(Spacer(1, 5))
        
        # Add market analysis header
        elements.append(RLParagraph("Market Analysis", styles['Heading1']))
        
        # Add project summaries
        elements.append(RLParagraph("Active Projects", styles['Heading2']))
        
        # Create project summary table
        project_data = [
            ['Project', 'Total Units', 'Sold', 'Standing', 'Avg PSF']
        ]
        
        for project in self.market_analyzer.project_data['active_projects']['projects']:
            project_data.append([
                project['name'],
                str(project['total_units']),
                str(project['sold_units']),
                str(project['standing_units']),
                f"${project.get('avg_psf', 'N/A'):,.2f}" if project.get('avg_psf') else 'N/A'
            ])
        
        # Create and add table
        table = Table(project_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Add market metrics
        elements.append(RLParagraph("Market Metrics", styles['Heading2']))
        market_metrics = self.market_analyzer.analysis_results['market_metrics']
        
        metrics_data = [
            ['Metric', 'Value'],
            ['Monthly Absorption Rate', f"{market_metrics['absorption_rate']*100:.1f}%"],
            ['Months of Supply', f"{market_metrics['months_of_supply']:.1f}"],
            ['Market Average PSF', f"${market_metrics['avg_psf']:,.2f}"],
            ['Price Elasticity', f"{market_metrics['price_elasticity']:.2f}"]
        ]
        
        metrics_table = Table(metrics_data)
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(metrics_table)
        
        return elements

    def _create_macro_landscape_pdf(self, styles) -> List:
        """Create macro landscape section for PDF"""
        elements = []
        
        # Title
        elements.append(RLParagraph("Macro Landscape", styles['Heading1']))
        elements.append(Spacer(1, 20))
        
        # Get macro data from market analysis
        market_factors = self.market_analyzer.analysis_results.get('market_factors', {})
        
        # 1. Interest Rates Section
        elements.append(RLParagraph("Interest Rate Trends", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        interest_rates = market_factors.get('interest_rates', {})
        current_rate = interest_rates.get('current', {}).get('rates', {}).get('5yr_fixed', 0)
        historical = interest_rates.get('historical_trends', {}).get('5yr_fixed', {})
        
        rate_data = [
            ['Period', '5-Year Fixed Rate'],
            ['Current Rate', f"{current_rate:.2f}%"],
            ['2024 YTD Average', f"{historical.get('2024_avg', 0):.2f}%"],
            ['2023 Average', f"{historical.get('2023_avg', 0):.2f}%"],
            ['2022 Average', f"{historical.get('2022_avg', 0):.2f}%"]
        ]
        
        table = Table(rate_data, colWidths=[200, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # 2. Employment Section
        elements.append(RLParagraph("Employment Metrics", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        employment = market_factors.get('employment', {})
        current_emp = employment.get('current_statistics', {})
        emp_trends = employment.get('historical_trends', {}).get('2024_ytd', {}).get('employment_rate', {})
        
        emp_data = [
            ['Metric', 'Value'],
            ['Current Employment Rate', f"{current_emp.get('employment_rate', 0):.1%}"],
            ['Current Unemployment Rate', f"{current_emp.get('unemployment_rate', 0):.1%}"],
            ['YTD Average', f"{emp_trends.get('average', 0):.1%}"],
            ['Trend', emp_trends.get('trend', 'N/A')]
        ]
        
        table = Table(emp_data, colWidths=[200, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # 3. Demographics Section
        elements.append(RLParagraph("Demographic Trends", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        demographics = market_factors.get('demographics', {}).get('household_income', {})
        current_income = demographics.get('current', {})
        growth = demographics.get('growth_metrics', {}).get('average_income_growth', {})
        
        demo_data = [
            ['Metric', 'Value'],
            ['Average Household Income', f"${current_income.get('average_income', 0):,.0f}"],
            ['Median Household Income', f"${current_income.get('median_income', 0):,.0f}"],
            ['Income Growth (YoY)', f"{growth.get('percentage', 0):.1f}%"],
            ['5-Year CAGR', f"{growth.get('cagr', 0):.1f}%"]
        ]
        
        table = Table(demo_data, colWidths=[200, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        
        return elements

    def _create_market_sensitivity_pdf(self, styles) -> List:
        """Create market sensitivity analysis section for PDF"""
        elements = []
        
        # Title
        elements.append(RLParagraph("Market Sensitivity Analysis", styles['Heading1']))
        elements.append(Spacer(1, 20))
        
        # Get actual calculated elasticity
        elasticity = self.market_analyzer.analysis_results['pricing_analysis'].get('price_elasticity', -0.8)
        
        # Add pricing sensitivity section directly (not using Excel worksheet)
        elements.append(RLParagraph("Pricing Sensitivity Analysis", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        sensitivity_data = [
            ['Metric', 'Value'],
            ['Price Elasticity', f"{elasticity:.2f}"],
            ['Interpretation', f"{abs(elasticity):.1f}% absorption change per 1% price change"],
            ['Data Source', "Calculated from historical project performance"],
            ['Confidence Level', "Based on actual market data"]
        ]
        
        # Create table for sensitivity data
        table = Table(sensitivity_data, colWidths=[200, 300])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Rest of the sensitivity analysis...
        # [Previous heatmap code remains unchanged]
        
        return elements

    def _add_pricing_sensitivity_section(self, ws, row: int) -> int:
        """Add pricing sensitivity section to Excel worksheet"""
        ws[f'A{row}'] = 'Pricing Sensitivity Analysis'
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 2
        
        # Get actual calculated elasticity
        elasticity = self.market_analyzer.analysis_results['pricing_analysis'].get('price_elasticity', -0.8)
        
        sensitivity_data = {
            'Price Elasticity': f"{elasticity:.2f}",
            'Interpretation': f"{abs(elasticity):.1f}% absorption change per 1% price change",
            'Data Source': "Calculated from historical project performance",
            'Confidence Level': "Based on actual market data"
        }
        
        self._add_metric_group(ws, row, 'Sensitivity Metrics', sensitivity_data)
        
        return row + len(sensitivity_data) + 3

    def _create_sensitivity_heatmap_table(self, data: List[List], min_val: float, max_val: float) -> Table:
        """Create a color-coded table to represent a heatmap"""
        def get_color(value: str) -> Color:
            """Get color based on value"""
            try:
                value = float(value.strip('%'))
                if value >= 65:  # Green zone
                    return colors.Color(0.2, 0.8, 0.2)  # Bright green
                elif value >= 45:  # Yellow zone
                    return colors.Color(0.9, 0.9, 0.2)  # Yellow
                else:  # Red zone
                    return colors.Color(0.8, 0.2, 0.2)  # Red
            except:
                return colors.white
        
        # Create table style
        style = [
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ]
        
        # Add color-coding for data cells
        for i in range(1, len(data)):
            for j in range(1, len(data[i])):
                style.append(('BACKGROUND', (j, i), (j, i), get_color(data[i][j])))
        
        table = Table(data, colWidths=[80] + [60]*(len(data[0])-1))
        table.setStyle(TableStyle(style))
        return table

    def _create_pdf_styles(self) -> Dict:
        """Create styles for PDF report"""
        styles = getSampleStyleSheet()
        
        # Modify existing styles
        styles['Normal'].fontName = 'Helvetica'
        styles['Normal'].fontSize = 10
        styles['Normal'].leading = 14
        
        # Modify existing heading styles
        styles['Heading1'].fontName = 'Helvetica-Bold'
        styles['Heading1'].fontSize = 16
        styles['Heading1'].leading = 20
        styles['Heading1'].spaceAfter = 12
        styles['Heading1'].textColor = colors.HexColor('#000000')
        
        styles['Heading2'].fontName = 'Helvetica-Bold'
        styles['Heading2'].fontSize = 14
        styles['Heading2'].leading = 18
        styles['Heading2'].spaceAfter = 10
        styles['Heading2'].textColor = colors.HexColor('#333333')
        
        styles['Heading3'].fontName = 'Helvetica-Bold'
        styles['Heading3'].fontSize = 12
        styles['Heading3'].leading = 16
        styles['Heading3'].spaceAfter = 8
        styles['Heading3'].textColor = colors.HexColor('#444444')
        
        # Modify existing bullet style
        styles['Bullet'].fontSize = 10
        styles['Bullet'].leading = 14
        styles['Bullet'].leftIndent = 20
        styles['Bullet'].bulletIndent = 10
        styles['Bullet'].spaceBefore = 2
        styles['Bullet'].spaceAfter = 2
        
        # Add custom styles
        custom_styles = {
            'TableBody': {
                'parent': 'Normal',
                'fontSize': 9,
                'leading': 12,
                'alignment': 1
            },
            'TableHeader': {
                'parent': 'TableBody',
                'fontName': 'Helvetica-Bold',
                'textColor': colors.whitesmoke,
                'backColor': colors.grey
            },
            'Caption': {
                'parent': 'Normal',
                'fontSize': 9,
                'leading': 12,
                'textColor': colors.HexColor('#666666'),
                'alignment': 1
            },
            'Note': {
                'parent': 'Normal',
                'fontSize': 8,
                'leading': 10,
                'textColor': colors.HexColor('#666666'),
                'italics': True
            },
            'Strategy': {
                'parent': 'Normal',
                'fontSize': 11,
                'leading': 14,
                'fontName': 'Helvetica-Bold',
                'textColor': colors.HexColor('#1a1a1a'),
                'spaceBefore': 6,
                'spaceAfter': 6,
                'leftIndent': 20
            }
        }
        
        # Add custom styles that don't exist in base stylesheet
        for style_name, style_props in custom_styles.items():
            if style_name not in styles:
                styles.add(ParagraphStyle(
                    name=style_name,
                    parent=styles[style_props['parent']],
                    **{k: v for k, v in style_props.items() if k != 'parent'}
                ))
        
        return styles

    def _create_unit_sensitivity_pdf(self, styles) -> List:
        """Create unit sensitivity analysis section for PDF"""
        elements = []
        
        # Title
        elements.append(RLParagraph("Unit Type Sensitivity Analysis", styles['Heading1']))
        elements.append(Spacer(1, 20))
        
        # Get unit type data
        unit_analysis = self.market_analyzer.analysis_results['unit_type_analysis']
        
        # 1. Unit Type Performance Matrix
        elements.append(RLParagraph("Unit Type Performance Matrix", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        performance_data = [
            ['Unit Type', 'Monthly Abs.', 'Annual Abs.', 'Avg PSF', 'Price Premium'],
        ]
        
        for unit_type in ['studios', 'one_bed', 'two_bed', 'three_bed']:
            metrics = unit_analysis[unit_type]
            performance_data.append([
                unit_type.replace('_', ' ').title(),
                f"{metrics['inventory_metrics']['absorption_rate']['monthly']:.1f}%",
                f"{metrics['inventory_metrics']['absorption_rate']['annualized']:.1f}%",
                f"${metrics['pricing_metrics']['avg_psf']:.2f}",
                f"{metrics['performance_metrics']['price_premium']:+.1f}%"
            ])
        
        table = Table(performance_data, colWidths=[100, 80, 80, 80, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # 2. Price Sensitivity by Unit Type
        elements.append(RLParagraph("Price Sensitivity by Unit Type", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        # Create price sensitivity matrix
        price_changes = [-10, -5, 0, 5, 10]
        sensitivity_data = [['Unit Type'] + [f"{change:+d}%" for change in price_changes]]
        
        for unit_type in ['studios', 'one_bed', 'two_bed', 'three_bed']:
            metrics = unit_analysis[unit_type]
            base_absorption = metrics['inventory_metrics']['absorption_rate']['monthly']
            
            row = [unit_type.replace('_', ' ').title()]
            for change in price_changes:
                # Calculate impact using price elasticity (-3.71 from market analysis)
                elasticity = -3.71
                impact = 1 + (change/100 * elasticity)
                new_absorption = base_absorption * impact
                row.append(f"{new_absorption:.1f}%")
            
            sensitivity_data.append(row)
        
        table = Table(sensitivity_data, colWidths=[100] + [60]*len(price_changes))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # 3. Unit Mix Recommendations
        elements.append(RLParagraph("Unit Mix Recommendations", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        recommendations = [
            ['Unit Type', 'Mix %', 'Target PSF', 'Strategy'],
            ['Studios', '9.0%', '$1,202.25', 'Value position for investors'],
            ['One Bed', '54.3%', '$1,145.00', 'Market anchor position'],
            ['Two Bed', '31.9%', '$1,087.75', 'End-user focus'],
            ['Three Bed', '4.8%', '$1,030.50', 'Limited release strategy']
        ]
        
        table = Table(recommendations, colWidths=[100, 60, 80, 180])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        
        return elements

    def _create_revenue_analysis_pdf(self, styles, period: str) -> List:
        """Create revenue analysis section for PDF"""
        elements = []
        
        # Title mapping and absorption targets
        absorption_targets = {
            '3_month': 0.50,    # 50% absorption
            '12_month': 0.65,   # 65% absorption
            '24_month': 0.825,  # 82.5% absorption
            '36_month': 1.00    # 100% absorption
        }
        
        title_text = {
            '3_month': 'Revenue Analysis - 3 Month (50% Absorption)',
            '12_month': 'Revenue Analysis - 12 Month (65% Absorption)',
            '24_month': 'Revenue Analysis - 24 Month (82.5% Absorption)',
            '36_month': 'Revenue Analysis - 36 Month (100% Absorption)'
        }
        
        if period not in absorption_targets:
            return elements  # Skip invalid periods
        
        elements.append(RLParagraph(title_text[period], styles['Heading1']))
        elements.append(Spacer(1, 20))
        
        # Unit type data
        unit_data = {
            'studios': {'units': 34, 'avg_size': 379, 'gross_psf': 1265.53, 'net_psf': 1202.25},
            'one_bed': {'units': 204, 'avg_size': 474, 'gross_psf': 1205.26, 'net_psf': 1145.00},
            'two_bed': {'units': 120, 'avg_size': 795, 'gross_psf': 1151.06, 'net_psf': 1087.75},
            'three_bed': {'units': 18, 'avg_size': 941, 'gross_psf': 1096.28, 'net_psf': 1030.50}
        }
        
        # Gross Revenue Analysis
        elements.append(RLParagraph("Gross Revenue Analysis", styles['Heading2']))
        elements.append(Spacer(1, 40))
        
        # Create Gross Revenue table
        gross_revenue_data = [
            ['Unit Type', 'Total Units', 'Absorbed Units', 'Avg Size', 'Gross PSF', 'Gross Price', 'Total Revenue']
        ]
        
        total_absorbed = 0
        total_gross_revenue = 0
        
        for unit_type, data in unit_data.items():
            absorbed_units = int(data['units'] * absorption_targets[period])
            gross_price = data['gross_psf'] * data['avg_size']
            gross_revenue = gross_price * absorbed_units
            
            total_absorbed += absorbed_units
            total_gross_revenue += gross_revenue
            
            gross_revenue_data.append([
                unit_type.replace('_', ' ').title(),
                str(data['units']),
                str(absorbed_units),
                str(data['avg_size']),
                f"${data['gross_psf']:,.2f}",
                f"${gross_price:,.0f}",
                f"${gross_revenue:,.0f}"
            ])
        
        # Add total row
        gross_revenue_data.append([
            'Total',
            '376',
            str(total_absorbed),
            '-',
            '-',
            '-',
            f"${total_gross_revenue:,.0f}"
        ])
        
        gross_table = Table(gross_revenue_data, [
            1.0*inch,  # Unit Type
            0.7*inch,  # Total Units
            0.8*inch,  # Absorbed Units
            0.7*inch,  # Avg Size
            0.8*inch,  # Gross PSF
            1.0*inch,  # Gross Price
            1.2*inch   # Total Revenue
        ])
        
        gross_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT')
        ]))
        
        elements.append(gross_table)
        elements.append(Spacer(1, 50))
        
        # Net Revenue Analysis
        elements.append(RLParagraph("Net Revenue Analysis", styles['Heading2']))
        elements.append(Spacer(1, 40))
        
        net_revenue_data = [
            ['Unit Type', 'Total Units', 'Absorbed Units', 'Avg Size', 'Net PSF', 'Net Price', 'Total Revenue']
        ]
        
        total_net_revenue = 0
        
        for unit_type, data in unit_data.items():
            absorbed_units = int(data['units'] * absorption_targets[period])
            net_price = data['net_psf'] * data['avg_size']
            net_revenue = net_price * absorbed_units
            
            total_net_revenue += net_revenue
            
            net_revenue_data.append([
                unit_type.replace('_', ' ').title(),
                str(data['units']),
                str(absorbed_units),
                str(data['avg_size']),
                f"${data['net_psf']:,.2f}",
                f"${net_price:,.0f}",
                f"${net_revenue:,.0f}"
            ])
        
        # Add total row
        net_revenue_data.append([
            'Total',
            '376',
            str(total_absorbed),
            '-',
            '-',
            '-',
            f"${total_net_revenue:,.0f}"
        ])
        
        net_table = Table(net_revenue_data, [
            1.0*inch,  # Unit Type
            0.7*inch,  # Total Units
            0.8*inch,  # Absorbed Units
            0.7*inch,  # Avg Size
            0.8*inch,  # Net PSF
            1.0*inch,  # Net Price
            1.2*inch   # Total Revenue
        ])
        
        net_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT')
        ]))
        
        elements.append(net_table)
        elements.append(Spacer(1, 50))
        
        # Monthly Absorption and Pricing Strategy
        elements.append(RLParagraph("Monthly Absorption and Pricing Strategy", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        # Create header rows
        monthly_data = [
            ['Month', 'Target %', 'Cumulative %',
             'Studios', '', '', 
             'One Bed', '', '',
             'Two Bed', '', '',
             'Three Bed', '', ''],
            ['', '', '',
             'Gross PSF', 'Incentive', 'Net PSF',
             'Gross PSF', 'Incentive', 'Net PSF',
             'Gross PSF', 'Incentive', 'Net PSF',
             'Gross PSF', 'Incentive', 'Net PSF']
        ]
        
        # Calculate monthly data
        cumulative = 0
        months = {'3_month': 3, '12_month': 12, '24_month': 24, '36_month': 36}
        num_months = months.get(period, 12)
        
        for month in range(num_months):
            if month < 3:  # First quarter
                target = 16.67
            elif month < 12:  # Rest of first year
                target = 1.67
            else:  # Beyond first year
                target = 1.46
                
            cumulative += target
            
            # Calculate incentives based on month
            if month < 3:
                studio_incentive = 5.0
                one_bed_incentive = 5.0
                two_bed_incentive = 5.5
                three_bed_incentive = 6.0
            elif month < 12:
                studio_incentive = 5.5
                one_bed_incentive = 5.5
                two_bed_incentive = 6.0
                three_bed_incentive = 6.5
            else:
                studio_incentive = 6.0
                one_bed_incentive = 6.0
                two_bed_incentive = 6.5
                three_bed_incentive = 7.0
            
            # Fixed gross PSF values
            studios_gross = 1265.53
            one_bed_gross = 1205.26
            two_bed_gross = 1151.06
            three_bed_gross = 1096.28
            
            monthly_data.append([
                f"Month {month + 1}",
                f"{target:.2f}%",
                f"{cumulative:.2f}%",
                f"${studios_gross:.2f}",
                f"{studio_incentive:.1f}%",
                f"${studios_gross * (1-studio_incentive/100):.2f}",
                f"${one_bed_gross:.2f}",
                f"{one_bed_incentive:.1f}%",
                f"${one_bed_gross * (1-one_bed_incentive/100):.2f}",
                f"${two_bed_gross:.2f}",
                f"{two_bed_incentive:.1f}%",
                f"${two_bed_gross * (1-two_bed_incentive/100):.2f}",
                f"${three_bed_gross:.2f}",
                f"{three_bed_incentive:.1f}%",
                f"${three_bed_gross * (1-three_bed_incentive/100):.2f}"
            ])
        
        # Create table with proper column widths
        table = Table(monthly_data, [
            0.6*inch,  # Month
            0.6*inch,  # Target %
            0.7*inch,  # Cumulative %
            0.7*inch, 0.5*inch, 0.7*inch,  # Studios
            0.7*inch, 0.5*inch, 0.7*inch,  # One Bed
            0.7*inch, 0.5*inch, 0.7*inch,  # Two Bed
            0.7*inch, 0.5*inch, 0.7*inch   # Three Bed
        ], repeatRows=2)
        
        # Style the table
        table.setStyle(TableStyle([
            # Header styling
            ('SPAN', (0,0), (0,1)),  # Month
            ('SPAN', (1,0), (1,1)),  # Target %
            ('SPAN', (2,0), (2,1)),  # Cumulative %
            ('SPAN', (3,0), (5,0)),  # Studios
            ('SPAN', (6,0), (8,0)),  # One Bed
            ('SPAN', (9,0), (11,0)), # Two Bed
            ('SPAN', (12,0), (14,0)), # Three Bed
            # Colors and borders
            ('BACKGROUND', (0,0), (-1,1), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,1), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            # Alternate row colors
            ('ROWBACKGROUNDS', (0,2), (-1,-1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(table)
        # elements.append(PageBreak())  # Remove this line
        
        return elements

    def _generate_sensitivity_commentary(self, unit_type: str, price_premium: float) -> Dict:
        """Generate sensitivity commentary for unit type"""
        # Get rate sensitivity values from market analyzer's calculations
        sensitivity_data = self.market_analyzer._calculate_unit_type_sensitivity(unit_type)
        
        # Use the actual calculated values for commentary
        rate_matrix = sensitivity_data.get('interest_rates_vs_price_change', {}).get('matrix', [[]])
        base_row_index = len(rate_matrix) // 2  # Middle row represents base case
        base_absorption = rate_matrix[base_row_index][len(rate_matrix[0]) // 2]  # Center value
        
        # Calculate actual rate sensitivity from the matrix
        rate_cut_absorption = rate_matrix[base_row_index - 1][len(rate_matrix[0]) // 2]  # One step down = rate cut
        rate_sensitivity = ((rate_cut_absorption - base_absorption) / base_absorption) * 100
        
        return {
            'Price Sensitivity': self._get_price_sensitivity_finding(unit_type),
            'Rate Sensitivity': (
                f"Historical data shows {abs(rate_sensitivity):.1f}% absorption increase "
                f"per 0.5% rate cut (2019-2023 data)"
            ),
            'Competitive Position': self._get_competitive_position_finding(unit_type, price_premium)
        }

    def _get_price_sensitivity_finding(self, unit_type: str) -> str:
        """Get price sensitivity finding for unit type"""
        sensitivities = {
            'studios': 17,
            'one_bed': 13,
            'two_bed': 8,
            'three_bed': 6
        }
        impact = sensitivities.get(unit_type, 10)
        return f"{'Highly' if impact > 15 else 'Moderate' if impact > 10 else 'Lower'} price sensitivity (±{impact}% impact)"

    def _generate_unit_sensitivity_findings(self, unit_type: str, metrics: Dict) -> Dict:
        """Generate sensitivity findings for unit type"""
        try:
            # Get base absorption rate
            base_absorption = metrics['inventory_metrics']['absorption_rate']['monthly']
            price_premium = metrics['performance_metrics'].get('price_premium', 0)
            
            # Generate findings using sensitivity commentary
            findings = self._generate_sensitivity_commentary(unit_type, price_premium)
            
            # Return findings as dictionary of key-value pairs
            return findings
                
        except Exception as e:
            print(f"Error generating unit sensitivity findings: {str(e)}")
            return {
                'Price Sensitivity': "Data not available",
                'Rate Sensitivity': "Data not available", 
                'Competitive Position': "Data not available"
            }